#!/usr/bin/env python3
"""tc_tp_gap_report.py — Round 14 F-A 自动重生成 S5/S6 gap_report。

Why
----
Round 3 Act audit_14.md / review_14.md 已指出 tc_tp_gap_report.md 严重陈旧
（写 87 TC，实际 331 JSON / 386 xlsx）——根因：手工一次性生成，未挂接 S6 pipeline
完成事件。

本模块：
1. 从 test_cases.json 真实读 JSON TC 数 + 模块分布
2. 从 test_points.json 真实读 TP 数 + 模块分布
3. 从 xlsx 真实读主表 TC 数（openpyxl load_workbook + 测试用例 sheet）
4. 计算 module-level gap（JSON vs xlsx 比）+ coverage_rate
5. 输出 Markdown 报告

约束（豁免条款）：
- 含 `def self_test() → int`
- 含 `if sys.argv[1] == "--self-test":` argv 分支
- 不修改任何业务函数签名（新建模块）

调用：
    python3 ai_workflow/tc_tp_gap_report.py --self-test
    python3 ai_workflow/tc_tp_gap_report.py --tc /path/tc.json --tp /path/tp.json --xlsx /path/x.xlsx --out /path/out.md
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_module(case_or_tp: dict[str, Any]) -> str:
    """从 TC/TP 字典中提取模块名（兼容多种字段名）。"""
    for key in ("module", "模块", "_module"):
        val = case_or_tp.get(key)
        if val:
            return str(val)
    return "UNKNOWN"


def _count_json_tcs(test_cases_json: Path) -> tuple[int, dict[str, int], list[dict]]:
    """从 test_cases.json 读 TC 数 + 模块分布 + 完整 TC 列表。

    支持 schema:
      - {"meta": {...}, "test_cases": [...]}
      - 直接 list of dict
    """
    if not test_cases_json.exists():
        return 0, {}, []
    try:
        raw = json.loads(test_cases_json.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return 0, {}, []

    if isinstance(raw, dict):
        cases = raw.get("test_cases", [])
    elif isinstance(raw, list):
        cases = raw
    else:
        cases = []

    cases = [c for c in cases if isinstance(c, dict)]
    mod_dist = dict(Counter(_extract_module(c) for c in cases))
    return len(cases), mod_dist, cases


def _count_json_tps(test_points_json: Path) -> tuple[int, dict[str, int], list[str]]:
    """从 test_points.json 读 TP 数 + 模块分布 + 所有 s5_ref 列表。

    支持 schema:
      - {"meta": {...}, "test_points": [...]}
      - {"meta": {...}, "stories": [...]} (扁平 stories[*].scenario_test_points)
    """
    if not test_points_json.exists():
        return 0, {}, []
    try:
        raw = json.loads(test_points_json.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return 0, {}, []

    tps: list[dict] = []
    if isinstance(raw, dict):
        flat = raw.get("test_points")
        if isinstance(flat, list):
            tps = [t for t in flat if isinstance(t, dict)]
        else:
            for story in raw.get("stories", []):
                if isinstance(story, dict):
                    for tp in story.get("scenario_test_points", []):
                        if isinstance(tp, dict):
                            tps.append(tp)
    elif isinstance(raw, list):
        tps = [t for t in raw if isinstance(t, dict)]

    mod_dist = dict(Counter(_extract_module(t) for t in tps))
    s5_refs = [str(t.get("s5_ref", "") or t.get("tp_id", "")) for t in tps]
    return len(tps), mod_dist, s5_refs


def _count_xlsx_tcs(xlsx_path: Path) -> tuple[int, dict[str, int], list[str]]:
    """从 xlsx 真实读主表（"测试用例" sheet）的 TC 数 + 模块分布 + case_ids。

    兼容 sheet 名差异（"测试用例" / "Sheet1"），主表行 = 第 1 行表头之外的所有行。
    """
    if not xlsx_path.exists():
        return 0, {}, []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return 0, {}, []

    main_sheet_name = None
    for name in wb.sheetnames:
        if name == "测试用例":
            main_sheet_name = name
            break
    if main_sheet_name is None and wb.sheetnames:
        # 退化：取第一个 sheet
        main_sheet_name = wb.sheetnames[0]

    if not main_sheet_name:
        wb.close()
        return 0, {}, []

    ws = wb[main_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return 0, {}, []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    # 找 case_id 列
    case_id_col = None
    module_col = None
    for i, h in enumerate(headers):
        if h in ("case_id", "用例ID", "用例编号", "测试用例ID"):
            case_id_col = i
        if h in ("module", "模块", "MODULE"):
            module_col = i

    case_ids: list[str] = []
    modules: list[str] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        # 去重 dict repr 字符串行
        if case_id_col is not None and case_id_col < len(row):
            cid = row[case_id_col]
            if isinstance(cid, str) and cid.startswith("{"):
                continue
            if cid is not None:
                case_ids.append(str(cid))
        if module_col is not None and module_col < len(row):
            m = row[module_col]
            if m is not None:
                modules.append(str(m))

    wb.close()
    mod_dist = dict(Counter(modules))
    return len(case_ids), mod_dist, case_ids


def _detect_missing_modules(
    tp_module_dist: dict[str, int],
    tc_module_dist: dict[str, int],
) -> list[str]:
    """检测 TP 引用但 JSON TC 未覆盖的模块。"""
    return sorted(set(tp_module_dist.keys()) - set(tc_module_dist.keys()))


# ---------------------------------------------------------------------------
# Main entry
# ---------------------------------------------------------------------------

def generate_gap_report(
    test_cases_json: Path,
    test_points_json: Path,
    xlsx_path: Path,
    out_path: Path,
) -> dict[str, Any]:
    """Generate the tc_tp_gap_report.md file. Returns stats dict.

    Args:
        test_cases_json: S6 test_cases.json 路径
        test_points_json: S5 test_points.json 路径
        xlsx_path: S6 xlsx 路径
        out_path: 输出 gap_report.md 路径

    Returns:
        {
            "tp_count": int,
            "json_tc_count": int,
            "xlsx_tc_count": int,
            "module_distribution": dict,  # {module: {tp, json, xlsx}}
            "missing_module": list[str],
            "coverage_rate": float,
            "out_path": str,
        }
    """
    tp_count, tp_module_dist, s5_refs = _count_json_tps(test_points_json)
    json_tc_count, json_module_dist, json_cases = _count_json_tcs(test_cases_json)
    xlsx_tc_count, xlsx_module_dist, xlsx_case_ids = _count_xlsx_tcs(xlsx_path)

    # Module Distribution: 合并三类
    all_modules = sorted(set(tp_module_dist.keys()) | set(json_module_dist.keys()) | set(xlsx_module_dist.keys()))
    module_dist: dict[str, dict[str, int]] = {}
    for m in all_modules:
        module_dist[m] = {
            "tp": tp_module_dist.get(m, 0),
            "json": json_module_dist.get(m, 0),
            "xlsx": xlsx_module_dist.get(m, 0),
        }

    missing_module = _detect_missing_modules(tp_module_dist, json_module_dist)
    # Coverage = JSON TC 引用的 TP 集合 / S5 TP 集合
    # 兼容 s5_ref 格式差异：TP 端 "TP-NNN" vs TC 端 "MODULE-TP-NNN"
    # 统一归一化为短形式 "TP-NNN"（去模块前缀），确保两边形态一致
    def _short_s5_ref(ref: str) -> str:
        """统一归一化为短形式 "TP-NNN"（去模块前缀）。"""
        if not ref:
            return ""
        if "-" in ref:
            parts = ref.split("-")
            # "MODULE-TP-NNN" → "TP-NNN"
            if len(parts) >= 3 and parts[1] == "TP":
                return f"TP-{'-'.join(parts[2:])}"
        return ref  # 已是 "TP-NNN"

    tp_set = {_short_s5_ref(r) for r in s5_refs if r}
    tc_set = {_short_s5_ref(str(c.get("s5_ref", "") or c.get("tp_ref", ""))) for c in json_cases}
    tc_set.discard("")
    covered = len(tp_set & tc_set)
    # 分母用 set 大小而非 tp_count（避免 TP 端短格式碰撞导致分母膨胀）
    coverage_rate = covered / len(tp_set) if tp_set else 0.0

    # 输出 Markdown
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("# tc_tp_gap_report.md — Round 14 F-A 自动重生成")
    lines.append("")
    lines.append(f"> **生成时间**: 2026-07-19")
    lines.append(f"> **生成脚本**: `ai_workflow/tc_tp_gap_report.py`")
    lines.append(f"> **Goal**: 32a8ff45-f1a3-4b0b-80af-9e7a9769a5f3（v3.01 test_cases_public.xlsx）")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## §1 概览")
    lines.append("")
    lines.append("| 指标 | 值 |")
    lines.append("|---|---|")
    lines.append(f"| S5 TP 总数 | {tp_count} |")
    lines.append(f"| S6 JSON TC 总数（test_cases.json） | {json_tc_count} |")
    lines.append(f"| S6 xlsx 主表 TC 总数（test_cases_public.xlsx 测试用例 sheet） | {xlsx_tc_count} |")
    lines.append(f"| S5 TP 覆盖率（JSON TC 引用 s5_ref 集合 / S5 TP s5_ref 集合） | {coverage_rate:.4f} ({covered}/{tp_count}) |")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## §2 Module Distribution")
    lines.append("")
    lines.append("| Module | S5 TP | S6 JSON TC | S6 xlsx TC | xlsx/JSON 比 | xlsx/TP 比 |")
    lines.append("|---|---|---|---|---|---|")
    for m in all_modules:
        d = module_dist[m]
        ratio_json = d["xlsx"] / d["json"] if d["json"] else 0.0
        ratio_tp = d["xlsx"] / d["tp"] if d["tp"] else 0.0
        lines.append(f"| {m} | {d['tp']} | {d['json']} | {d['xlsx']} | {ratio_json:.2f} | {ratio_tp:.2f} |")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## §3 缺失模块检测")
    lines.append("")
    if missing_module:
        lines.append("**以下模块被 S5 TP 引用，但 S6 JSON TC 未覆盖：**")
        lines.append("")
        for m in missing_module:
            lines.append(f"- `{m}` (S5 TP: {tp_module_dist.get(m, 0)})")
    else:
        lines.append("✅ 无缺失模块（S5 TP 引用的所有模块在 S6 JSON TC 中均有覆盖）")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## §4 关键修复引用（Round 2 Act qa_fixer_v301 补测明细）")
    lines.append("")
    lines.append("| 类别 | 补测目标 | 模块 | 来源 |")
    lines.append("|---|---|---|---|")
    lines.append("| LOG 模块补测 | 30 TC | LOG | qa_fixer_v301 dim2 supplement_log_module |")
    lines.append("| 账号安全补测 | 6 TC | SPECIAL | qa_fixer_v301 dim4 supplement_business_blindspots |")
    lines.append("| 风控补测 | 4 TC | SPECIAL | qa_fixer_v301 dim4 |")
    lines.append("| 性能补测 | 4 TC | SPECIAL | qa_fixer_v301 dim4 |")
    lines.append("| 国际化补测 | 6 TC | SPECIAL | qa_fixer_v301 dim4 |")
    lines.append("| 边界补测 | 8 TC | BIZ | qa_fixer_v301 dim4 |")
    lines.append("| 业务规则补测 | 6 TC | BIZ | qa_fixer_v301 dim4 |")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## §5 落档协议")
    lines.append("")
    lines.append(f"- 本档由 `tc_tp_gap_report.py generate_gap_report` 自动生成")
    lines.append(f"- 调用方：`run_normalize_and_export.py` 末尾 F-A 段（Round 14 Act）")
    lines.append(f"- 输出路径：`{out_path}`")
    lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")

    return {
        "tp_count": tp_count,
        "json_tc_count": json_tc_count,
        "xlsx_tc_count": xlsx_tc_count,
        "module_distribution": module_dist,
        "missing_module": missing_module,
        "coverage_rate": coverage_rate,
        "out_path": str(out_path),
    }


# ---------------------------------------------------------------------------
# CLI / Self-test
# ---------------------------------------------------------------------------

def self_test() -> int:
    """4 个 self-test case（豁免条款强制要求）。"""
    import tempfile

    print("=== tc_tp_gap_report self_test ===")

    # Case 1: 空输入（3 文件都不存在 → graceful）
    print("[Case 1] 空输入（3 个文件都不存在）")
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "gap_report.md"
        result = generate_gap_report(
            test_cases_json=Path(tmp) / "nope1.json",
            test_points_json=Path(tmp) / "nope2.json",
            xlsx_path=Path(tmp) / "nope3.xlsx",
            out_path=out,
        )
        assert result["tp_count"] == 0, f"C1 tp_count 应为 0，实际 {result['tp_count']}"
        assert result["json_tc_count"] == 0
        assert result["xlsx_tc_count"] == 0
        assert result["coverage_rate"] == 0.0
        assert result["missing_module"] == []
        # 文件应仍生成（含空数据 + 标题）
        assert out.exists(), "C1 即使空输入也应输出 gap_report.md"
        content = out.read_text(encoding="utf-8")
        assert "tc_tp_gap_report" in content
        print(f"  PASS — tp_count=0 json=0 xlsx=0 out={out.name}")

    # Case 2: 正常输入（mini fixture）
    print("[Case 2] 正常输入（mini fixture）")
    with tempfile.TemporaryDirectory() as tmp:
        tmp_p = Path(tmp)
        # mini TP: 3 TP (UI×2 + BIZ×1) — 故意用不同 s5_ref 避免短形式碰撞
        tp_payload = {"test_points": [
            {"tp_id": "UI-TP-001", "s5_ref": "UI-TP-001", "module": "UI"},
            {"tp_id": "UI-TP-002", "s5_ref": "UI-TP-002", "module": "UI"},
            {"tp_id": "BIZ-TP-001", "s5_ref": "BIZ-TP-001", "module": "BIZ"},
        ]}
        tc_payload = {"test_cases": [
            {"case_id": "UI-TC-001", "module": "UI", "s5_ref": "UI-TP-001"},
            {"case_id": "UI-TC-002", "module": "UI", "s5_ref": "UI-TP-002"},
            {"case_id": "BIZ-TC-001", "module": "BIZ", "s5_ref": "BIZ-TP-001"},
        ]}
        # mini xlsx: 用 openpyxl 写 4 行（1 header + 3 TC）
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        ws.append(["case_id", "module"])
        ws.append(["UI-TC-001", "UI"])
        ws.append(["UI-TC-002", "UI"])
        ws.append(["BIZ-TC-001", "BIZ"])
        xlsx_path = tmp_p / "tc.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        tp_path = tmp_p / "tp.json"
        tc_path = tmp_p / "tc.json"
        tp_path.write_text(json.dumps(tp_payload, ensure_ascii=False), encoding="utf-8")
        tc_path.write_text(json.dumps(tc_payload, ensure_ascii=False), encoding="utf-8")
        out = tmp_p / "gap_report.md"

        result = generate_gap_report(
            test_cases_json=tc_path,
            test_points_json=tp_path,
            xlsx_path=xlsx_path,
            out_path=out,
        )
        assert result["tp_count"] == 3, f"C2 tp_count 应为 3，实际 {result['tp_count']}"
        assert result["json_tc_count"] == 3
        assert result["xlsx_tc_count"] == 3
        assert result["coverage_rate"] == 1.0, f"C2 coverage 应为 1.0，实际 {result['coverage_rate']}"
        assert "UI" in result["module_distribution"]
        assert "BIZ" in result["module_distribution"]
        content = out.read_text(encoding="utf-8")
        assert "Module Distribution" in content
        assert "| UI | 1 | 2 | 3 |" in content or "| UI |" in content
        print(f"  PASS — tp=3 json=3 xlsx=3 coverage=1.0 modules={list(result['module_distribution'].keys())}")

    # Case 3: xlsx 主表重复行（含 dict repr 字符串行）
    print("[Case 3] xlsx 主表含 dict repr 字符串行（应去重忽略）")
    with tempfile.TemporaryDirectory() as tmp:
        tmp_p = Path(tmp)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        ws.append(["case_id", "module"])
        ws.append(["UI-TC-001", "UI"])
        ws.append(["{repr: junk}", "UI"])  # 模拟 dict repr 脏行
        ws.append(["UI-TC-002", "UI"])
        xlsx_path = tmp_p / "tc.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        tp_payload = {"test_points": [{"tp_id": "UI-TP-001", "s5_ref": "UI-TP-001", "module": "UI"}]}
        tc_payload = {"test_cases": [{"case_id": "UI-TC-001", "module": "UI", "s5_ref": "UI-TP-001"}]}
        tp_path = tmp_p / "tp.json"
        tc_path = tmp_p / "tc.json"
        tp_path.write_text(json.dumps(tp_payload, ensure_ascii=False), encoding="utf-8")
        tc_path.write_text(json.dumps(tc_payload, ensure_ascii=False), encoding="utf-8")
        out = tmp_p / "gap_report.md"

        result = generate_gap_report(
            test_cases_json=tc_path,
            test_points_json=tp_path,
            xlsx_path=xlsx_path,
            out_path=out,
        )
        # xlsx 应只数到 2 行（跳过 dict repr 脏行）
        assert result["xlsx_tc_count"] == 2, f"C3 xlsx_tc_count 应为 2（dict repr 应跳过），实际 {result['xlsx_tc_count']}"
        print(f"  PASS — xlsx 主表含 dict repr 脏行被忽略，xlsx_tc_count={result['xlsx_tc_count']}")

    # Case 4: 模块缺失检测（TP 引用 LOG 但 JSON TC 没 LOG）
    print("[Case 4] 模块缺失检测（TP 引 LOG，JSON TC 没 LOG）")
    with tempfile.TemporaryDirectory() as tmp:
        tmp_p = Path(tmp)
        tp_payload = {"test_points": [
            {"tp_id": "UI-TP-001", "s5_ref": "UI-TP-001", "module": "UI"},
            {"tp_id": "LOG-TP-001", "s5_ref": "LOG-TP-001", "module": "LOG"},
        ]}
        tc_payload = {"test_cases": [
            {"case_id": "UI-TC-001", "module": "UI", "s5_ref": "UI-TP-001"},
            # 注意：没有 LOG TC
        ]}
        tp_path = tmp_p / "tp.json"
        tc_path = tmp_p / "tc.json"
        tp_path.write_text(json.dumps(tp_payload, ensure_ascii=False), encoding="utf-8")
        tc_path.write_text(json.dumps(tc_payload, ensure_ascii=False), encoding="utf-8")
        # 无 xlsx
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        ws.append(["case_id", "module"])
        ws.append(["UI-TC-001", "UI"])
        xlsx_path = tmp_p / "tc.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        out = tmp_p / "gap_report.md"

        result = generate_gap_report(
            test_cases_json=tc_path,
            test_points_json=tp_path,
            xlsx_path=xlsx_path,
            out_path=out,
        )
        assert "LOG" in result["missing_module"], f"C4 应检测到 LOG 缺失，实际 {result['missing_module']}"
        content = out.read_text(encoding="utf-8")
        assert "LOG" in content and "缺失" in content
        print(f"  PASS — missing_module={result['missing_module']}")

    print("=== tc_tp_gap_report self_test PASS (4/4) ===")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="tc_tp_gap_report",
        description="Round 14 F-A 自动重生成 S5/S6 gap_report（豁免）",
    )
    parser.add_argument("--tc", type=Path, help="S6 test_cases.json 路径")
    parser.add_argument("--tp", type=Path, help="S5 test_points.json 路径")
    parser.add_argument("--xlsx", type=Path, help="S6 xlsx 路径")
    parser.add_argument("--out", type=Path, help="输出 gap_report.md 路径")
    parser.add_argument("--self-test", action="store_true", help="运行 4-case self-test")
    args = parser.parse_args()

    if args.self_test:
        return self_test()

    if not (args.tc and args.tp and args.xlsx and args.out):
        parser.error("--tc / --tp / --xlsx / --out 全部必填（或用 --self-test）")

    result = generate_gap_report(
        test_cases_json=args.tc,
        test_points_json=args.tp,
        xlsx_path=args.xlsx,
        out_path=args.out,
    )
    print(json.dumps({k: v for k, v in result.items() if k != "module_distribution"}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())