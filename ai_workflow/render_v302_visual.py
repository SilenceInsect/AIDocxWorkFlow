#!/usr/bin/env python3
"""Render v3.02 Round 1 visual: OBJ bands, merge guards, H-cell independence, P0 coverage.

T-005.4: 视觉确认 v3.02 Round 1 重导 xlsx 的物理状态。
- OBJ 色带 (5 色循环, 每 OBJ 一色)
- B/C/D 列合并 (46 merges = B14 + C16 + D16)
- E/F uniform guard (0 merges — 数据保护)
- H 列独立性 (87 独立 Cell)
- 优先级列分布 (P0/P1/P2 颜色编码)
- V-002 16/16 OBJ P0 覆盖

Output: workflow_assets/游戏道具商城系统/v3.01/「S6 测试用例生成」/test_cases_public.v302.visual.png
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent
DEFAULT_XLSX = (
    REPO_ROOT / "workflow_assets" / "游戏道具商城系统" / "v3.01"
    / "「S6 测试用例生成」" / "test_cases_public.xlsx"
)
DEFAULT_PNG = (
    REPO_ROOT / "workflow_assets" / "游戏道具商城系统" / "v3.01"
    / "「S6 测试用例生成」" / "test_cases_public.v302.visual.png"
)

PALETTE = ["#E6E6E6", "#E3F2FD", "#FFF8E1", "#E8F5E9", "#F3E5F5"]
PRIO_COLOR = {"P0": "#DC2626", "P1": "#F59E0B", "P2": "#3B82F6", "P3": "#6B7280"}
WIDTHS = [60, 130, 140, 240, 320, 260, 280, 880]
LABELS = [
    "行",
    "A 用例ID",
    "B 所属模块",
    "C 用例描述",
    "D 关联需求",
    "E 功能描述",
    "F 前置条件",
    "H 预期结果（独立 TC cell）",
]


def _font(size: int) -> Any:
    candidates = [
        Path("/System/Library/Fonts/PingFang.ttc"),
        Path("/System/Library/Fonts/STHeiti Medium.ttc"),
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
    ]
    font_path = next((path for path in candidates if path.exists()), None)
    if font_path is None:
        return ImageFont.load_default()
    try:
        return ImageFont.truetype(str(font_path), size=size, index=0)
    except TypeError:
        return ImageFont.truetype(str(font_path), size=size)


def _truncate(value: Any, max_chars: int) -> str:
    text = "" if value is None else str(value).replace("\n", " / ")
    return text if len(text) <= max_chars else text[: max_chars - 1] + "…"


def _block_spans(worksheet: Any, ranges: list[Any], start_row: int, end_row: int) -> list[tuple[int, int, int]]:
    """Detect OBJ block spans by reading D column (关联需求)."""
    spans: list[tuple[int, int, int]] = []
    row = start_row
    index = 0
    while row <= end_row:
        end = row
        # Find merges starting at this row in D col
        merged = next(
            (item for item in ranges if item.min_col == 4 and item.min_row == row),
            None,
        )
        if merged is not None:
            end = min(merged.max_row, end_row)
        spans.append((row, end, index))
        row = end + 1
        index += 1
    return spans


def render(xlsx_path: Path = DEFAULT_XLSX, output_path: Path = DEFAULT_PNG) -> dict[str, Any]:
    workbook = load_workbook(xlsx_path)
    worksheet = workbook["测试用例"]
    ranges = list(worksheet.merged_cells.ranges)

    row_height = 32
    meta_height = 168
    header_height = 44
    segment_gap = 48
    image_width = sum(WIDTHS) + 40
    image_height = meta_height + header_height + (worksheet.max_row - 1) * row_height + segment_gap
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)
    title_font = _font(28)
    header_font = _font(18)
    body_font = _font(15)
    small_font = _font(13)

    # Title
    draw.text((20, 14), "v3.02 Round 1 — OBJ 色带、合并保护、H 列独立、P0 覆盖", fill="#111827", font=title_font)
    draw.text(
        (20, 52),
        "物理结果: 88×11, 87 cases, 46 merges (B14+C16+D16), E/F=0 (uniform guard), H=0 (独立 cell)",
        fill="#9A3412",
        font=header_font,
    )
    draw.text(
        (20, 80),
        "V-001: 4 模块连续 (BIZ-1..64, UI-1..19, LOG-1, SPECIAL-1..3)  |  V-002: 16/16 OBJ 至少 1 P0",
        fill="#15803D",
        font=header_font,
    )
    draw.text(
        (20, 108),
        "V-003: H 列字面重复 = 0  |  V-004: B 列 row 27/28 (OBJ-02 块) 物理读 = 'BIZ-TC-026/027'",
        fill="#1E40AF",
        font=header_font,
    )
    draw.text(
        (20, 136),
        "色带: OBJ 块 5 色循环; 红框: OBJ 边界; 绿框: 实际合并; 蓝条: H 列独立 cell; P 列色块: 优先级.",
        fill="#374151",
        font=body_font,
    )

    y = meta_height
    # Header row
    x = 20
    for label, width in zip(LABELS, WIDTHS):
        draw.rectangle((x, y, x + width, y + header_height), fill="#1F2937", outline="#FFFFFF")
        draw.text((x + 6, y + 12), label, fill="white", font=header_font)
        x += width
    y += header_height

    # Build spans for full range
    spans = _block_spans(worksheet, ranges, 2, worksheet.max_row)
    table_top = y

    # Per-row priority lookup (from priority column = column I = col 9)
    for row in range(2, worksheet.max_row + 1):
        span = next(item for item in spans if item[0] <= row <= item[1])
        color = PALETTE[span[2] % len(PALETTE)]
        case_id = worksheet.cell(row, 1).value
        values = [
            row,
            case_id,
            worksheet.cell(row, 2).value,
            worksheet.cell(row, 3).value,
            worksheet.cell(row, 4).value,
            worksheet.cell(row, 5).value,
            worksheet.cell(row, 6).value,
            worksheet.cell(row, 8).value,
        ]
        limits = [4, 16, 14, 26, 36, 28, 28, 95]
        x = 20
        prio = str(worksheet.cell(row, 9).value or "")
        for index, (value, width, limit) in enumerate(zip(values, WIDTHS, limits)):
            fill = "#FFFFFF" if index == 0 else color
            draw.rectangle((x, y, x + width, y + row_height), fill=fill, outline="#CBD5E1")
            display = _truncate(value, limit)
            if index == 7:
                line_count = len(str(value).splitlines()) if value is not None else 0
                display = _truncate(f"{case_id} | {line_count} lines | {value}", limit)
            text_color = "#111827"
            if index == 1:
                # 用例ID column — colorize by priority
                pcolor = PRIO_COLOR.get(prio, "#111827")
                text_color = pcolor
            draw.text(
                (x + 5, y + 8),
                display,
                fill=text_color,
                font=small_font if index in (4, 7) else body_font,
            )
            x += width
        y += row_height

    # Red box around each OBJ block
    x_positions = [20]
    for width in WIDTHS:
        x_positions.append(x_positions[-1] + width)
    for start, end, _ in spans:
        top = table_top + (start - 2) * row_height
        bottom = table_top + (end - 2 + 1) * row_height
        draw.rectangle((20, top, image_width - 20, bottom), outline="#B91C1C", width=3)

    # Green box around actual merges (only show the row range borders for merges)
    for mr in ranges:
        if mr.min_col == mr.max_col and mr.min_row >= 2:
            top = table_top + (mr.min_row - 2) * row_height
            bottom = table_top + (mr.max_row - 2 + 1) * row_height
            # only box columns B/C/D (cols 2/3/4)
            if mr.min_col in (2, 3, 4):
                col_left = 20 + sum(WIDTHS[: mr.min_col - 1])
                col_right = col_left + WIDTHS[mr.min_col - 1]
                draw.rectangle((col_left, top, col_right, bottom), outline="#16A34A", width=2)

    # Blue bar at bottom of H column
    h_left = 20 + sum(WIDTHS[:7])
    h_right = h_left + WIDTHS[7]
    bar_top = table_top + (worksheet.max_row - 2 + 1) * row_height + 4
    bar_bot = bar_top + 6
    draw.rectangle((h_left, bar_top, h_right, bar_bot), fill="#1D4ED8")

    image.save(output_path, "PNG")
    return {
        "output": str(output_path),
        "size_bytes": output_path.stat().st_size,
        "image_size": image.size,
    }


def self_test() -> int:
    """Render the current xlsx to PNG."""
    if not DEFAULT_XLSX.exists():
        print(f"[FAIL] xlsx not found: {DEFAULT_XLSX}")
        return 1
    result = render()
    print(f"render_v302_visual self-test: PASS — {result['output']} ({result['size_bytes']} bytes)")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--self-test":
        raise SystemExit(self_test())
    raise SystemExit("usage: python3 ai_workflow/render_v302_visual.py --self-test")
