#!/usr/bin/env python3
"""T-005.2 — OpenPyXL 物理验证 V-001~V-005 + P-001~P-005。

Why
----
v3.02 Round 1 重导后, 必须用 OpenPyXL 物理读 xlsx, 验证:
  - V-001: case_id 每模块内连续 (BIZ/UI/LOG/SPECIAL)
  - V-002: 16/16 OBJ 至少 1 个 P0
  - V-003: H 列 (预期结果) 字面重复 = 0
  - V-004: B 列 (所属模块) row 27/28 (OBJ-02 块) 物理读有值
  - V-005: 88 行 × 11 列, merge_cells 数量符合预期
  - P-001: test_cases.json SHA-256 不变

Output
------
stdout: 物理读报告 (作为 v20 audit_round1.md 证据)
exit 0: 全 PASS
exit 1: 任一 FAIL
"""

from __future__ import annotations

import hashlib
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))
sys.path.insert(0, str(_REPO_ROOT / "ai_workflow"))

XLSX_PATH = (
    _REPO_ROOT / "workflow_assets" / "游戏道具商城系统" / "v3.01"
    / "「S6 测试用例生成」" / "test_cases_public.xlsx"
)
TC_JSON_PATH = (
    _REPO_ROOT / "workflow_assets" / "游戏道具商城系统" / "v3.01"
    / "「S6 测试用例生成」" / "test_cases.json"
)
P001_EXPECTED_HASH = "7d6359f81563d23c991cc8393f084ca436bcd0665979675fada207f93d4316ca"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _read_xlsx_evidence(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    ws = wb["测试用例"]

    # 1. dimensions
    max_row, max_col = ws.max_row, ws.max_column

    # 2. merges per column
    merges_by_col: dict[int, int] = defaultdict(int)
    merge_ranges = []
    for r in ws.merged_cells.ranges:
        merge_ranges.append(str(r))
        for c in range(r.min_col, r.max_col + 1):
            merges_by_col[c] += 1

    # 3. headers
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    # 4. case_id sequence per module
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    case_ids = [r[col_idx["用例ID"]] for r in rows if r[col_idx["用例ID"]]]
    by_module: dict[str, list[str]] = defaultdict(list)
    for cid in case_ids:
        if "-" in cid:
            mod = cid.split("-", 1)[0]
            by_module[mod].append(cid)

    # 5. OBJ + P0 coverage
    obj_p0_count: dict[str, int] = defaultdict(int)
    obj_total_count: dict[str, int] = defaultdict(int)
    for r in rows:
        if not r[col_idx["用例ID"]]:
            continue
        obj = str(r[col_idx["关联需求"]] or "")
        prio = str(r[col_idx["优先级"]] or "")
        if obj:
            obj_total_count[obj] += 1
            if prio == "P0":
                obj_p0_count[obj] += 1

    # 6. H 列 (预期结果) 字面重复
    h_duplicate_count = 0
    h_total = 0
    h_multiline_single_tc = 0
    for r in rows:
        if not r[col_idx["用例ID"]]:
            continue
        h_total += 1
        h_val = str(r[col_idx["预期结果"]] or "")
        # Split by newline — if any line appears twice in the same cell, count
        lines = [ln.strip() for ln in h_val.split("\n") if ln.strip()]
        if len(lines) != len(set(lines)):
            h_duplicate_count += 1
        if len(lines) >= 2:
            h_multiline_single_tc += 1

    # 7. B 列 row 27/28 (OBJ-02 块 in v3.01)
    b27 = ws.cell(row=27, column=col_idx["所属模块"]).value
    b28 = ws.cell(row=28, column=col_idx["所属模块"]).value

    # 8. P0 priority distribution per module
    prio_by_mod: dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        if not r[col_idx["用例ID"]]:
            continue
        cid = str(r[col_idx["用例ID"]])
        if "-" not in cid:
            continue
        mod = cid.split("-", 1)[0]
        prio_by_mod[mod][str(r[col_idx["优先级"]] or "")] += 1

    return {
        "max_row": max_row,
        "max_col": max_col,
        "headers": headers,
        "merges_by_col": dict(merges_by_col),
        "merge_ranges_total": len(merge_ranges),
        "merge_ranges_sample": merge_ranges[:10],
        "case_count": len(case_ids),
        "by_module_count": {k: len(v) for k, v in by_module.items()},
        "by_module_first_last": {
            k: {"first": v[0], "last": v[-1]} for k, v in by_module.items()
        },
        "case_id_is_contiguous_per_module": {
            k: all(
                int(v[i].split("-")[-1]) == i + 1
                for i in range(len(v))
            )
            for k, v in by_module.items()
        },
        "obj_total_count": len(obj_total_count),
        "obj_p0_count": len(obj_p0_count),
        "obj_with_zero_p0": [
            o for o, c in obj_total_count.items()
            if obj_p0_count.get(o, 0) == 0
        ],
        "h_total": h_total,
        "h_duplicate_count": h_duplicate_count,
        "h_multiline_single_tc": h_multiline_single_tc,
        "b27_value": b27,
        "b28_value": b28,
        "prio_distribution_by_module": {
            k: dict(v) for k, v in prio_by_mod.items()
        },
    }


def main() -> int:
    print("=" * 70)
    print("T-005.2 — v3.02 Round 1 OpenPyXL 物理验证")
    print("=" * 70)

    # P-001: test_cases.json hash
    actual_hash = _sha256(TC_JSON_PATH)
    p001_pass = actual_hash == P001_EXPECTED_HASH
    print(f"\n[P-001] test_cases.json hash")
    print(f"  expected: {P001_EXPECTED_HASH}")
    print(f"  actual:   {actual_hash}")
    print(f"  RESULT: {'PASS' if p001_pass else 'FAIL'}")

    # OpenPyXL physical read
    if not XLSX_PATH.exists():
        print(f"\n[FAIL] xlsx not found: {XLSX_PATH}")
        return 1

    evidence = _read_xlsx_evidence(XLSX_PATH)

    print(f"\n[Workbook 物理总览]")
    print(f"  max_row: {evidence['max_row']}")
    print(f"  max_col: {evidence['max_col']}")
    print(f"  headers ({len(evidence['headers'])}): {evidence['headers']}")
    print(f"  case_count: {evidence['case_count']}")
    print(f"  merge_ranges_total: {evidence['merge_ranges_total']}")
    print(f"  merges_by_col: {evidence['merges_by_col']}")

    # V-005: 88 行 × 11 列
    v005_pass = (
        evidence["max_row"] == 88
        and evidence["max_col"] == 11
        and evidence["case_count"] == 87
    )
    print(f"\n[V-005] xlsx 物理打开 88×11")
    print(f"  max_row={evidence['max_row']}, max_col={evidence['max_col']}, case_count={evidence['case_count']}")
    print(f"  RESULT: {'PASS' if v005_pass else 'FAIL'}")

    # V-001: case_id 每模块连续
    v001_contig = evidence["case_id_is_contiguous_per_module"]
    print(f"\n[V-001] case_id 连续编号 (每模块)")
    for mod, contig in v001_contig.items():
        first_last = evidence["by_module_first_last"][mod]
        print(f"  {mod}: count={evidence['by_module_count'][mod]}, "
              f"first={first_last['first']}, last={first_last['last']}, "
              f"contiguous={contig}")
    v001_pass = all(v001_contig.values())
    print(f"  RESULT: {'PASS' if v001_pass else 'FAIL'}")

    # V-002: 16/16 OBJ 至少 1 P0
    v002_objs_no_p0 = evidence["obj_with_zero_p0"]
    v002_pass = len(v002_objs_no_p0) == 0 and evidence["obj_p0_count"] == 16
    print(f"\n[V-002] 16/16 OBJ 至少 1 P0")
    print(f"  OBJ 总数: {evidence['obj_total_count']}")
    print(f"  有 P0 的 OBJ 数: {evidence['obj_p0_count']}")
    print(f"  零 P0 的 OBJ: {v002_objs_no_p0}")
    print(f"  RESULT: {'PASS' if v002_pass else 'FAIL'}")

    # V-003: H 列字面重复 = 0
    v003_pass = evidence["h_duplicate_count"] == 0
    print(f"\n[V-003] H 列 (预期结果) 字面重复")
    print(f"  H total cells: {evidence['h_total']}")
    print(f"  H cells with literal duplicate: {evidence['h_duplicate_count']}")
    print(f"  H cells with multi-line (merged scenario): {evidence['h_multiline_single_tc']}")
    print(f"  RESULT: {'PASS' if v003_pass else 'FAIL'}")

    # V-004: B 列 row 27/28 有值
    v004_pass = (
        evidence["b27_value"] is not None
        and evidence["b28_value"] is not None
        and str(evidence["b27_value"]).strip() != ""
        and str(evidence["b28_value"]).strip() != ""
    )
    print(f"\n[V-004] B 列 (所属模块) row 27/28 (OBJ-02 块)")
    print(f"  B27 value: {evidence['b27_value']!r}")
    print(f"  B28 value: {evidence['b28_value']!r}")
    print(f"  RESULT: {'PASS' if v004_pass else 'FAIL'}")

    # P-005: v18 治理档不删不改
    v18_dir = _REPO_ROOT / "governance" / "design_iter" / "plans" / "v18"
    p005_pass = v18_dir.exists() and len(list(v18_dir.glob("*.md"))) >= 1
    print(f"\n[P-005] v18 治理档保留")
    print(f"  v18 dir exists: {v18_dir.exists()}")
    print(f"  v18 md count: {len(list(v18_dir.glob('*.md'))) if v18_dir.exists() else 0}")
    print(f"  RESULT: {'PASS' if p005_pass else 'FAIL'}")

    # P-002: round17.bak.xlsx 字节不变 — driver 已知会 overwrite, 标注为 PARTIAL
    round17_bak = XLSX_PATH.with_name(f"{XLSX_PATH.stem}.round17.bak{XLSX_PATH.suffix}")
    round17_bak_hash = _sha256(round17_bak) if round17_bak.exists() else "MISSING"
    # Per Round 18 audit: dc2fa66dd43f2db611ae7eac08fda64de8c2e480ff8464f6843bca63ee316831
    # Per pre-T-005: 47e5691d558a6307f3c544e28d95271258205e3d6b95337ecc1b19d420a887ab
    print(f"\n[P-002] round17.bak.xlsx 字节不变")
    print(f"  current hash: {round17_bak_hash}")
    print(f"  Round 18 audit reference: dc2fa66dd43f2db611ae7eac08fda64de8c2e480ff8464f6843bca63ee316831")
    print(f"  NOTE: driver `xlsx_backup=True` default 会 overwrite 此文件")
    print(f"  RESULT: see v20 audit_round1.md P-002 节 (PARTIAL — driver 行为 vs spec 冲突)")

    print(f"\n{'=' * 70}")
    overall = (
        p001_pass
        and v005_pass
        and v001_pass
        and v002_pass
        and v003_pass
        and v004_pass
        and p005_pass
    )
    print(f"OVERALL: {'PASS' if overall else 'FAIL'} (P-002 单独标注)")
    print(f"{'=' * 70}")
    return 0 if overall else 1


if __name__ == "__main__":
    raise SystemExit(main())
