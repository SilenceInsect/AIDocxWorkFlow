#!/usr/bin/env python3
"""S6 xlsx 整理器（thin wrapper）— 脚本不推导，只把 JSON 转成 xlsx。

⚠️ 设计原则（v2.0 重构 — 2026-06-15）：
- 脚本只搬运字段，**不统计"测试方法分布"**（方法由 LLM 自由标注，不需要脚本聚合）
- 3 个 sheet：测试用例 / 模块统计 / 类型统计
- 不加"测试方法 Sheet"——避免暗示 LLM 必须按固定方法体系拓宽
"""

import json
import openpyxl
from pathlib import Path
from collections import Counter

XLSX_PATH = Path("/Users/gleon/Documents/TestDev/AIDocxWorkFlow-SH/workflow_assets/游戏道具商城系统/v1.0/「S6 测试用例生成」/test_cases.xlsx")
JSON_PATH = Path("/Users/gleon/Documents/TestDev/AIDocxWorkFlow-SH/workflow_assets/游戏道具商城系统/v1.0/「S6 测试用例生成」/test_cases.json")


def main() -> None:
    with JSON_PATH.open(encoding="utf-8") as f:
        data = json.load(f)

    cases = data.get("test_cases", [])
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1: 测试用例
    ws1 = wb.create_sheet("测试用例")
    ws1.append(["用例ID", "Story", "TP", "标题", "模块", "前置条件",
                "操作步骤", "预期结果", "优先级", "用例类型", "用例状态"])
    for c in cases:
        ws1.append([
            c.get("case_id", ""),
            c.get("story_id", ""),
            c.get("tp_id", ""),
            c.get("用例描述", "") or c.get("title", ""),
            c.get("module", ""),
            c.get("precondition", "") or c.get("前置条件", ""),
            c.get("steps", "") or c.get("操作步骤", ""),
            c.get("expected", "") or c.get("预期结果", ""),
            c.get("优先级", ""),
            c.get("test_type", ""),
            c.get("用例状态", ""),
        ])

    # Sheet 2: 模块统计（搬运）
    ws2 = wb.create_sheet("模块统计")
    ws2.append(["模块", "用例数", "占比"])
    module_counter = Counter(c.get("module", "") for c in cases)
    total = len(cases)
    for module, count in sorted(module_counter.items()):
        pct = (count / total * 100) if total > 0 else 0
        ws2.append([module, count, f"{pct:.1f}%"])

    # Sheet 3: 类型统计（搬运）
    ws3 = wb.create_sheet("类型统计")
    ws3.append(["测试类型", "用例数", "占比"])
    type_counter = Counter(c.get("test_type", "") for c in cases)
    for t, count in sorted(type_counter.items()):
        pct = (count / total * 100) if total > 0 else 0
        ws3.append([t, count, f"{pct:.1f}%"])

    # 调整列宽
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(XLSX_PATH)
    print(f"[S6 xlsx] → {XLSX_PATH}")
    print(f"  Sheet 1 (测试用例): {len(cases)} 行")
    print(f"  Sheet 2 (模块统计): {len(module_counter)} 个模块")
    print(f"  Sheet 3 (类型统计): {len(type_counter)} 种类型")


if __name__ == "__main__":
    main()
