#!/usr/bin/env python3
"""Round 15 主调：同源用例聚合 + L1∧L2 + 双 Sheet 重导出 xlsx。

Why
----
v3.01 test_cases.json 全部 331 条 TC 都是单步单预期。同 obj+fp+scenario 被拆成 6~11 条
独立 TC，下游测试工程师无法使用。本脚本在 Round 12 既有基础设施之上加 1 步聚合：

    load v3.01 json (read-only)
      → normalize (case_id + bilingual alias)
      → merge_grouped (同源聚合；NEW)
      → L1∧L2 评估
      → _save_xlsx (双 Sheet 分流)

只动 xlsx 产物；v3.01 test_cases.json 永远 read-only（out_of_scope §10）。
"""

from __future__ import annotations

import copy
import json
import shutil
import sys
from collections.abc import Mapping
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
if str(_REPO_ROOT / "ai_workflow") not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT / "ai_workflow"))

from ai_workflow.case_id_and_field_normalizer import (
    evaluate_status,
    load_payload,
    normalize_payload,
    renumber_cases_per_module,
)
from ai_workflow.scenario_group_merger import merge_grouped_inplace
from ai_workflow.test_case_formatter import _ROUND17_XLSX_PROFILE, _save_xlsx


REQ_NAME = "游戏道具商城系统"
VERSION = "v3.01"
DEFAULT_S6_DIR = (
    _REPO_ROOT / "workflow_assets" / REQ_NAME / VERSION / "「S6 测试用例生成」"
)
DEFAULT_TC_JSON = DEFAULT_S6_DIR / "test_cases.json"
DEFAULT_XLSX_OUT = DEFAULT_S6_DIR / "test_cases_public.xlsx"
DEFAULT_TPS_JSON = (
    _REPO_ROOT / "workflow_assets" / REQ_NAME / VERSION / "「S5 测试点生成」" / "test_points.json"
)
DEFAULT_OBJS_JSON = (
    _REPO_ROOT / "workflow_assets" / REQ_NAME / VERSION / "「S2 需求拆解」" / "requirement_objects.json"
)


def _load_objs_tps(
    objs_path: Path | str | None,
    tps_path: Path | str | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    objs: list[dict[str, Any]] = []
    tps: list[dict[str, Any]] = []

    if objs_path and Path(objs_path).exists():
        try:
            raw = json.loads(Path(objs_path).read_text(encoding="utf-8"))
            if isinstance(raw, list):
                objs = [o for o in raw if isinstance(o, dict)]
            elif isinstance(raw, dict):
                objs_list = raw.get("requirement_objects") or raw.get("objs")
                if isinstance(objs_list, list):
                    objs = [o for o in objs_list if isinstance(o, dict)]
                elif isinstance(objs_list, dict):
                    objs = [objs_list]
        except Exception as e:  # pragma: no cover
            print(f"[warn] failed to read objs: {e}")

    if tps_path and Path(tps_path).exists():
        try:
            raw = json.loads(Path(tps_path).read_text(encoding="utf-8"))
            if isinstance(raw, list):
                tps = [t for t in raw if isinstance(t, dict)]
            elif isinstance(raw, dict):
                flat = raw.get("test_points") or raw.get("test_points_by_story")
                if isinstance(flat, list):
                    tps = [t for t in flat if isinstance(t, dict)]
                elif flat is None:
                    for story in raw.get("stories", []):
                        if isinstance(story, dict):
                            for tp in story.get("scenario_test_points", []):
                                if isinstance(tp, dict):
                                    tps.append(tp)
        except Exception as e:  # pragma: no cover
            print(f"[warn] failed to read tps: {e}")

    return objs, tps


def merge_and_export(
    *,
    src_json: Path | str = DEFAULT_TC_JSON,
    xlsx_out: Path | str = DEFAULT_XLSX_OUT,
    objs_json: Path | str | None = DEFAULT_OBJS_JSON,
    tps_json: Path | str | None = DEFAULT_TPS_JSON,
    xlsx_backup: bool = True,
    l2_mode: str = "lenient",
    apply_renumber: bool = True,
) -> dict[str, Any]:
    """Run the Round 15 pipeline (normalize → merge → renumber → L1∧L2 → xlsx).

    Returns a summary dict with before/after counts, l1/l2 results, writeback
    report, merge group sizes, renumber stats, and final xlsx stats.
    ``src_json`` is read-only — the renumber is purely an in-memory mutation
    (T-003 / V-001 fix; P-001 BLOCKER protected).

    ``apply_renumber`` defaults to True so callers get the v3.01 BLOCKER fix
    for free; set False to skip renumbering (e.g. when debugging other steps).
    """
    src = Path(src_json)
    out_xlsx = Path(xlsx_out)
    if not src.exists():
        raise FileNotFoundError(f"source json not found: {src}")

    payload = load_payload(src)
    cases_in: list[dict[str, Any]]
    if isinstance(payload, dict):
        cases_in = payload.get("test_cases") or []
    elif isinstance(payload, list):
        cases_in = payload
    else:
        cases_in = []

    input_count = len(cases_in)

    # Step 1: in-place normalize (case_id + bilingual alias + OBJ P0 coverage).
    # T-004 fix: route through `normalize_payload` instead of calling the lower
    # level primitives in a loop. The previous inline loop skipped
    # `enforce_obj_p0_coverage`, leaving 12/16 OBJs without P0 coverage
    # (V-002 BLOCKER not actually exercised at export time).
    cases_in, _counters = normalize_payload(payload if isinstance(payload, dict) else cases_in)

    # Step 2: same-scenario merge (NEW in Round 15)
    if isinstance(payload, dict):
        cases = merge_grouped_inplace(payload)
    else:
        cases = merge_grouped_inplace(cases_in)

    merged_count = len(cases)
    group_sizes = _group_size_stats(cases)

    # Step 2.5: per-module contiguous renumber (T-003 / V-001 BLOCKER).
    # Pure in-memory: src_json on disk is never touched. The xlsx downstream
    # sees gap-free numbering per module (e.g. BIZ-TC-001..027, UI-TC-001..N).
    renumber_stats = renumber_cases_per_module(cases, apply=apply_renumber)

    # Step 3: load S2 / S5 context for L1 traceability
    objs, tps = _load_objs_tps(
        Path(objs_json) if objs_json else None,
        Path(tps_json) if tps_json else None,
    )

    # Step 4: L1∧L2 evaluation + per-case writeback (lenient by default)
    eval_result = evaluate_status(cases, objs=objs, tps=tps, run_l2=True, l2_mode=l2_mode)

    # Step 5: preserve the previous round's workbook, then export.
    # Round 18 — backup name bumped from round16.bak to round17.bak.
    if xlsx_backup and out_xlsx.exists():
        backup_path = out_xlsx.with_name(f"{out_xlsx.stem}.round17.bak{out_xlsx.suffix}")
        shutil.copy2(out_xlsx, backup_path)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx(
        cases,
        out_xlsx.parent,
        export_profile=copy.deepcopy(_ROUND17_XLSX_PROFILE),
        output_path=out_xlsx,
        sort_options={
            "by": ["obj_id", "case_id"],
            "obj_color_band": True,
            # Round 18: extend merge_obj_headers to 5 columns (per user report
            # R18-B1). The physical merge is conditional — see
            # _populate_worksheet_with_obj_grouping + _is_column_uniform — so
            # non-uniform columns (e.g. E/F when they vary per FP) are skipped
            # to prevent data loss. 16/16 v3.01 OBJ blocks have non-uniform E/F,
            # so the actual merge count stays at 48 (= 16 OBJ × 3 uniform cols)
            # for this dataset.
            "merge_obj_headers": ["用例描述", "所属模块", "关联需求", "功能描述", "前置条件"],
        },
        description_index_sheet=True,
    )

    return {
        "src_json": str(src),
        "xlsx_out": str(out_xlsx),
        "input_cases": input_count,
        "merged_cases": merged_count,
        "group_size_distribution": group_sizes,
        "renumber": renumber_stats,
        "evaluation": eval_result,
        "l2_mode": l2_mode,
    }


def _group_size_stats(cases: list[dict[str, Any]]) -> dict[str, int]:
    """How many cases have 1 step / 2-3 steps / 4+ steps after merge."""
    stats = {"1_step": 0, "2_to_3_steps": 0, "4_plus_steps": 0, "max_steps_in_a_case": 0}
    for case in cases:
        if not isinstance(case, dict):
            continue
        n = len(case.get("steps") or [])
        if n == 1:
            stats["1_step"] += 1
        elif 2 <= n <= 3:
            stats["2_to_3_steps"] += 1
        elif n >= 4:
            stats["4_plus_steps"] += 1
        if n > stats["max_steps_in_a_case"]:
            stats["max_steps_in_a_case"] = n
    return stats


def _summary_for_stdout(summary: Mapping[str, Any]) -> dict[str, Any]:
    """Trim the payload to keys humans want to see."""
    writeback = summary["evaluation"]["writeback"]
    transitions = writeback.get("transition_count") or writeback.get("changed_cases", 0)
    return {
        "src_json": summary["src_json"],
        "xlsx_out": summary["xlsx_out"],
        "input_cases": summary["input_cases"],
        "merged_cases": summary["merged_cases"],
        "compression_ratio": (
            round(summary["input_cases"] / summary["merged_cases"], 3)
            if summary["merged_cases"] else None
        ),
        "group_size_distribution": summary["group_size_distribution"],
        # T-003 / V-001 BLOCKER fix evidence
        "renumber": {
            "rewrites": summary["renumber"].get("rewrites", 0),
            "already_canonical": summary["renumber"].get("already_canonical", False),
            "by_module": summary["renumber"].get("by_module", {}),
        },
        "l1_passed": summary["evaluation"]["l1_result"]["passed"],
        "l1_errors": summary["evaluation"]["l1_result"]["stats"],
        "l2_passed": summary["evaluation"]["l2_result"]["passed"],
        "l2_failed_count": summary["evaluation"]["l2_result"].get("failed_count", 0),
        "writeback_target": writeback.get("target_status"),
        "writeback_changed": transitions,
        "l2_mode": summary["l2_mode"],
    }


def self_test() -> int:
    """Cover end-to-end: merge → L1∧L2 → xlsx partition. Uses a tempdir."""
    import tempfile

    # Build a mini payload with 2 same-scenario pairs so we exercise merge
    # AND end-to-end L1∧L2 → dual sheet.
    from ai_workflow.run_normalize_and_export import _build_mini_payload

    base: dict[str, Any] = _build_mini_payload()
    # Round 15 requires test_scenario on every case for the merge key.
    # The legacy mini payload from Round 12 doesn't have it; we add it here.
    for tc in base["test_cases"]:
        tc.setdefault("test_scenario", tc.get("用例描述", "通用"))
        tc.setdefault("preconditions", ["初始化完成"])

    extra: dict[str, Any] = {
        **base["test_cases"][0],
        "case_id": "TC-1b",
        "steps": [{"step_num": 1, "action": "complete"}],
        "expected_results": ["completed"],
    }
    base["test_cases"].append(extra)

    with tempfile.TemporaryDirectory() as tmpdir:
        src = Path(tmpdir) / "tc.json"
        xlsx = Path(tmpdir) / "out.xlsx"
        src.write_text(json.dumps(base, ensure_ascii=False, indent=2), encoding="utf-8")

        summary = merge_and_export(
            src_json=src,
            xlsx_out=xlsx,
            objs_json=None,
            tps_json=None,
            xlsx_backup=False,
            # T-003 / V-001 — apply_renumber=False keeps this self-test focused on
            # the merge + xlsx partition contract. Without the flag the per-module
            # renumber sorts cases by obj_id and the TC-1 legacy id becomes
            # BIZ-TC-001 (since the merged TC-1's obj_id sorts before OBJ-MALL-001
            # in the test fixture), breaking the "find UI-TC-001" assertion. The
            # renumber behaviour itself is covered by case_id_and_field_normalizer
            # tests 12-15.
            apply_renumber=False,
        )

        # Input 6 cases (5 original + 1 duplicate scenario); merge should
        # collapse the duplicate scenario with case_id="TC-1b" into TC-1.
        assert summary["input_cases"] == 6, summary
        assert summary["merged_cases"] == 5, summary

        # Read-back the xlsx
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx))
        assert "测试用例" in wb.sheetnames, wb.sheetnames
        assert "Draft-Rejected附录" in wb.sheetnames, wb.sheetnames
        # Round 16 — new Sheet 2 「用例描述索引」 is opt-in via the merge_and_export
        # default; the self_test exercises the merge+export end-to-end so the
        # index sheet should be present.
        assert "用例描述索引" in wb.sheetnames, wb.sheetnames

        # The merged TC-1 case should now have 2 steps + 2 expected_results.
        ws = wb["测试用例"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        merged_pair = [r for r in rows if r[0] == "UI-TC-001"]
        assert len(merged_pair) == 1, merged_pair
        # The Round 17 profile inserts 关联需求 after 用例描述.
        merged_row = merged_pair[0]
        assert "1." in str(merged_row[6]) and "2." in str(merged_row[6]), merged_row[6]
        assert "ok" in str(merged_row[7]) and "completed" in str(merged_row[7]), (
            f"expected_results not merged: {merged_row[7]!r}"
        )

    print("run_round15_merge_export self-test: PASS")
    return 0


def main() -> dict[str, Any]:
    # P-002 protection: by default, do NOT overwrite round17.bak.xlsx on re-export.
    # The legacy `xlsx_backup=True` behaviour preserves the previous xlsx as a
    # round-17 snapshot — but in iterative loops that itself becomes the
    # backup-clobbering hazard. Set AIDOCX_ROUND15_XLSX_BACKUP=1 to force the
    # legacy copy-on-write behaviour (e.g. when you explicitly want a new
    # backup snapshot).
    import os
    backup = os.environ.get("AIDOCX_ROUND15_XLSX_BACKUP", "0") == "1"
    summary = merge_and_export(
        src_json=DEFAULT_TC_JSON,
        xlsx_out=DEFAULT_XLSX_OUT,
        objs_json=DEFAULT_OBJS_JSON,
        tps_json=DEFAULT_TPS_JSON,
        xlsx_backup=backup,
    )
    print(json.dumps(_summary_for_stdout(summary), ensure_ascii=False, indent=2))
    return summary


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--self-test":
        raise SystemExit(self_test())
    main()
