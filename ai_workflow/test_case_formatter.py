#!/usr/bin/env python3
"""S5/S6 自动化格式化模块。

- S5: compose_test_points_from_structure() — 从 backlog 自动生成测试点骨架
- S6: format_test_cases() — 自动分配用例ID、规范化字段，默认生成 .json；项目级导出再生成 .md/.xlsx
"""

from __future__ import annotations
import copy, json, re, zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from collections.abc import Mapping  # used by Round 16 helpers
from xml.sax.saxutils import escape

_ROOT = Path(__file__).parent.parent.resolve()
WF    = _ROOT / "workflow_assets"
PROJECT_KNOWLEDGE_ROOT = _ROOT / "knowledge" / "project_local"


# ─────────────────────────────────────────────────────────────────────────────
# S5: 测试点骨架生成
# ─────────────────────────────────────────────────────────────────────────────

def compose_test_points_from_structure(
    backlog_json: dict,
    req_name: str = "游戏道具商城系统",
    version: str = "v1.0",
) -> dict:
    """
    从 backlog.json 生成测试点结构骨架。

    骨架仅保留 Story 原始信息，不预填任何推导字段。
    `scenario_test_points[]` 完全由 LLM 按 SKILL.md §1.4 推理填入：
    - 4 类型枚举（POSITIVE/BOUNDARY/NEGATIVE/EXCEPTION）
    - 8 模块（UI/BIZ/CONFIG/UTIL/LINK/LOG/SPECIAL/HINT）
    - s4_reference、applies_rule、is_assumed 等必填字段

    返回:
        dict: {
            "version": str,
            "stage": "S5",
            "req_name": str,
            "stories": [
                {
                    "story_id": "BIZ-PURCHASE-001",
                    "title": "购买确认流程",
                    "module": "BIZ",
                    "precondition": "...",
                    "input_data": "...",
                    "expected_output": "...",
                    "acceptance_criteria": [...],
                    "scenario_test_points": [],   # LLM 推理填入
                }
            ]
        }
    """
    epics = backlog_json.get("epics", [])
    result = {
        "version": version,
        "stage": "S5",
        "date": datetime.now().strftime("%Y-%m-%d"),
        "req_name": req_name,
        "stories": [],
    }

    for epic in epics:
        epic_id   = epic["id"]
        epic_mod  = epic.get("module", epic_id.split("-")[0])
        for story in epic.get("stories", []):
            sid    = story["id"]
            title  = story["title"]
            acs    = story.get("acceptance_criteria", [])
            pre    = story.get("precondition", "")
            inp    = story.get("input_data", "")
            out    = story.get("expected_output", "")

            result["stories"].append({
                "story_id": sid,
                "epic_id": epic_id,
                "module": epic_mod,
                "title": title,
                "precondition": pre,
                "input_data": inp,
                "expected_output": out,
                "acceptance_criteria": acs,
                "scenario_test_points": [],   # LLM 按 SKILL.md §1.4 推理填入
            })

    return result




# ─────────────────────────────────────────────────────────────────────────────
# S6: 测试用例格式化
# ─────────────────────────────────────────────────────────────────────────────

def format_test_cases(
    ai_raw_output: str | list,
    breakdown: dict,
    test_points: dict,
    req_name: str = "游戏道具商城系统",
    version: str = "v1.0",
    project_name: str | None = None,
    output_dir: Path | str | None = None,
) -> dict:
    """
    将 AI 输出的用例内容规范化为完整格式。

    自动化处理：
    - 分配用例 ID（按模块序号：UI-TC-001, BIZ-TC-001...）
    - 规范化字段
    - 步骤编号
    - 去重
    - 始终生成公共 JSON + 公共级 XLSX（v35 强制约束）
    - 项目级 Markdown/XLSX 仅在已确认项目且存在项目级导出配置时生成

    参数:
        ai_raw_output: AI 生成的用例内容（字符串或列表）
        breakdown: backlog.json 数据
        test_points: S5 输出的测试点 JSON
        project_name: 已确认的项目名。公共级 xlsx 不受此参数影响，始终产出。
        output_dir: 输出目录（默认 <req_name>/v1.0/「S6 测试用例生成」/）
    """
    if output_dir is None:
        output_dir = WF / req_name / version / "「S6 测试用例生成」"
    else:
        output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        from ai_workflow.stage_context_builder import build_stage_context
        from ai_workflow.stage_gatekeeper import write_read_ack
        stage_context = build_stage_context("S6", req_name, version, project_name=project_name, persist=True)
        write_read_ack(stage_context, output_dir=output_dir)
    except Exception as exc:
        print(f"[S6] stage_context/read_ack 生成失败: {exc}")

    # 解析 AI 输出
    if isinstance(ai_raw_output, str):
        parsed = _parse_ai_output(ai_raw_output)
    else:
        parsed = ai_raw_output

    # 在生成 coverage ledger 前，尽量把用例回挂到 Story。
    parsed = _attach_story_metadata(parsed, breakdown, test_points)

    # 自动分配 ID
    cases = _assign_ids(parsed, test_points)

    # 构建最终 JSON
    result_json = {
        "version": version,
        "stage": "S6",
        "date": datetime.now().strftime("%Y-%m-%d"),
        "req_name": req_name,
        "summary": _build_summary(cases),
        "test_cases": cases,
    }

    # 保存 JSON
    json_path = output_dir / "test_cases.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(result_json, f, ensure_ascii=False, indent=2)

    from ai_workflow.case_status_writer import apply_l1_status
    from ai_workflow.validators.l1_s6 import L1S6Validator

    l1_result = L1S6Validator().run_l1_check(json_path)
    status_writeback = apply_l1_status(cases, l1_result)
    result_json["status_writeback"] = status_writeback
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(result_json, f, ensure_ascii=False, indent=2)
    print(f"[S6] JSON → {json_path}")
    print(f"[S6] L1 {'PASS' if l1_result['passed'] else 'FAIL'} → {status_writeback['target_status']}")

    export_profile = _load_project_export_profile(project_name)
    template_path = _load_project_xlsx_template(project_name)
    md_path = None
    xlsx_path = None
    xlsx_project_path = None

    # v35 强制约束：公共级 xlsx 必须始终产出，不受 project_name 影响
    xlsx_path = _save_xlsx(
        cases,
        output_dir,
        export_profile=copy.deepcopy(_DEFAULT_XLSX_PROFILE),
        template_path=None,  # 公共 xlsx 不使用项目模板
    )
    print(f"[S6] XLSX（公共级）→ {xlsx_path}")

    # 项目级 xlsx：仅在已确认项目且存在项目级 profile 时生成
    if export_profile:
        md_path = _save_md(cases, output_dir)
        print(f"[S6] Markdown → {md_path}")
        xlsx_project_path = _save_xlsx(
            cases,
            output_dir,
            export_profile=export_profile,
            template_path=template_path,
        )
        print(f"[S6] XLSX（项目级）→ {xlsx_project_path}")
    else:
        print("[S6] 未确认项目或未检测到项目级导出配置，仅生成公共级 xlsx")

    gate = None
    try:
        from ai_workflow.stage_gatekeeper import run_postflight_gate
        gate = run_postflight_gate("S6", req_name, version, project_name=project_name, persist=True)
    except Exception as exc:
        print(f"[S6] postflight gate 执行失败: {exc}")

    return {
        "json": str(json_path),
        "md": str(md_path) if md_path else None,
        "xlsx": str(xlsx_path),          # v35: 公共级 xlsx 始终有值
        "xlsx_project": str(xlsx_project_path) if xlsx_project_path else None,  # v35: 项目级 xlsx（可选）
        "case_count": len(cases),
        "summary": result_json["summary"],
        "gate": gate,
        "status_writeback": status_writeback,
        "export_profile": export_profile,
        "xlsx_template": str(template_path) if template_path else None,
        "project_name": project_name,
    }


# ── 子函数 ──────────────────────────────────────────────────────────────────

# 模块定义见 .cursor/MODULES.md（项目级唯一真相源）
# 本文件不重写模块表，仅引用；如模块调整，只改 MODULES.md
#
# 重要规则（与 MODULES.md §1 严格一致）：
# 1. HINT **不**作为 Epic ID 前缀——HINT 仅作为 TP module 字段的取值
# 2. 测试用例 case_id 的"模块"前缀必须从 8 模块中取全名（UI/BIZ/CONFIG/UTIL/LINK/LOG/SPECIAL/HINT）
# 3. 模块字段支持中英双语（"界面" = "UI"），但 case_id 前缀统一用英文全名
_MODULE_SEQ = ["UI", "BIZ", "CONFIG", "UTIL", "LINK", "LOG", "SPECIAL", "HINT"]
_MODULE_COUNTER = {m: 0 for m in _MODULE_SEQ}

# 中英模块名归一化映射（"界面" / "UI" / "ui" → "UI"）
# 用于 summary 分组时合并中英键
# 重要：同时支持 **v1.1 旧 4 字母缩写**（CFG/LNK/SPC/HNT）和 **8 模块全名**
_MODULE_NAME_NORMALIZE = {
    # ── 中文 → 英文全名 ──
    "界面":   "UI",
    "业务":   "BIZ",
    "配置":   "CONFIG",
    "辅助":   "UTIL",
    "关联":   "LINK",
    "日志":   "LOG",
    "特殊":   "SPECIAL",
    "提示":   "HINT",
    # ── 英文全名（pass-through，大小写不敏感）──
    "ui":   "UI",
    "biz":  "BIZ",
    "config": "CONFIG",
    "UTIL":  "UTIL",
    "link": "LINK",
    "log":  "LOG",
    "special": "SPECIAL",
    "hint": "HINT",
    "UI":   "UI",
    "BIZ":  "BIZ",
    "CONFIG": "CONFIG",
    "UTIL":  "UTIL",
    "LINK": "LINK",
    "LOG":  "LOG",
    "SPECIAL": "SPECIAL",
    "HINT": "HINT",
    # ── v1.1 旧 4 字母缩写 → 现行全名（v1.1 时 _module_prefix() 错误返回的缩写）──
    "cfg":  "CONFIG",   # 旧 "CFG" → CONFIG
    "lnk":  "LINK",     # 旧 "LNK" → LINK
    "spc":  "SPECIAL",  # 旧 "SPC" → SPECIAL
    "hnt":  "HINT",     # 旧 "HNT" → HINT
    "CFG":  "CONFIG",
    "LNK":  "LINK",
    "SPC":  "SPECIAL",
    "HNT":  "HINT",
}

# v1.1 → v1.2/v1.7 旧枚举升级映射（v1.1 旧数据加载时执行）
# 仅保留仍可识别的旧枚举；新枚举不在此表中
_V11_TO_CURRENT = {
    # UI v1.1 4 枚举 → v1.2 11 枚举
    "CONTROL_EXISTENCE":  "CONTROL_RENDER",   # 默认归 RENDER；人工判定时可拆出 STATE
    "INTERACTION":        "PURE_INTERACTION", # 默认归 PURE；UI 基础功能可拆出 CONTROL_BASE_FUNC
    "LAYOUT":             "LAYOUT_ADAPT",
    "RESOLUTION_COMPAT":  "LAYOUT_ADAPT",
    # UTIL v1.1 4 枚举 → v1.2 14 枚举
    "TOOL_UTIL":          "COMMON_UTIL",
    "NETWORK_LAYER":      "NETWORK_LAYER",
    "CACHE_HIT_RATE":     "CACHE_HIT_RATE",
    "RESOURCE_MGMT":      "RESOURCE_MGMT",
    # BIZ v1.1 4 枚举 → v1.2 9 枚举（ACTIVITY_OPEN_CLOSE 拆分）
    "ACTIVITY_OPEN_CLOSE": "BIZ_STATE_MACHINE",  # 状态流转类；业务结果→BIZ_LOGIC
    "PROTOCOL":            "BIZ_PROTOCOL",
    "DB_PERSIST":          "BIZ_DB_PERSIST",
    "ENTITY_CACHE":        "CACHE_HIT_RATE",     # v1.1 旧 BIZ 枚举 → v1.2 移交 UTIL
    # LINK v1.1 3 枚举 → v1.2 6 枚举
    "CORRELATION_TEST":    "INTERNAL_BIZ_LINKAGE",
    "REGRESSION_TEST":     "CROSS_SERVER_SYNC",  # 跨服部分；第三方部分→EXTERNAL_THIRD_PARTY
    "MULTI_TENANT_SYNC":   "MULTI_CLIENT_SYNC",  # 多端一致性
    # LOG v1.1 4 枚举 → v1.9 13 枚举
    "ASSET_CHANGE":        "LOG_ASSET_AUDIT",
    "PROGRESS_TRIGGER":    "LOG_MONITOR",
    "ANOMALY":             "LOG_CRASH_REPORT",
    "AUDIT_TRAIL":         "LOG_INTEGRITY",
    # SPECIAL v1.1 5 枚举 → v1.2 9 枚举
    "DUPLICATE_PACKET":    "BOUNDARY_EXTREME",   # 业务层去重；流量层→WEAK_NET_RATE_LIMIT
    "HIGH_FREQ_PACKET":    "BOUNDARY_EXTREME",   # 业务层高频；服务端 HA→SERVER_HA_RISK
    "WEAK_NETWORK":        "WEAK_NET_RATE_LIMIT",
    "SWITCH_TO_BACKGROUND": "BG_FG_SWITCH",
    "ANTI_CHEAT":          "ANTI_CHEAT",
    # HINT v1.6.1 6 枚举 → v1.7 13 枚举
    "RED_DOT":             "RED_DOT_BADGE",      # 语义升级：扩展为「红点+角标+数字」
    "ITEM_FLOAT":          "ITEM_FLOAT",
    "CURRENCY_FLOAT":      "CURRENCY_FLOAT",
    "MODAL_DIALOG":        "MODAL_DIALOG",
    "TOAST":               "TOAST",
    "FLOAT_NOTIFY":        "FLOAT_NOTIFY",
    # 旧/已废/凭空出现的枚举（v1.6.1 S6 test_cases.json 发现的）
    "SYS_MSG":             "MODAL_DIALOG",       # v1.6.1 凭空出现 → 迁移至 MODAL_DIALOG 或 TOAST
    "ENTITY_REL":          "BIZ_DATA_FLOW",      # 旧枚举（v1.1）→ BIZ 端服数据流
    "GAME_LOG":            "LOG_OPERATION",      # 旧枚举 → LOG 通用
}


def normalize_module_name(name: str) -> str | None:
    """归一化模块名（中文/英文缩写/全名/小写 → 8 模块全名）。

    规则：
    - 8 模块之一（大小写不敏感）→ 返回全名
    - 中文名（界面/业务/...）→ 返回对应全名
    - 旧 v1.1 缩写（CFG/LNK/SPC/HNT）→ 返回 CONFIG/LINK/SPECIAL/HINT
    - 不识别 → 返回 None（调用方应处理）

    Args:
        name: 任意形式的模块名

    Returns:
        8 模块之一（CONFIG/UI/BIZ/UTIL/LINK/LOG/SPECIAL/HINT），不识别返回 None
    """
    if not name:
        return None
    key = str(name).strip()
    return _MODULE_NAME_NORMALIZE.get(key) or _MODULE_NAME_NORMALIZE.get(key.upper())


def migrate_old_enum(old_value: str, fallback_module: str = "BIZ") -> tuple[str, str | None]:
    """将 v1.1 旧枚举值升级为 v1.2/v1.7 现行枚举。

    Args:
        old_value: 旧枚举字符串
        fallback_module: 当无法识别时，回退到该模块的默认枚举

    Returns:
        (new_enum, migration_note) - new_enum 是 v1.2+ 现行枚举；
        migration_note 是 None（无迁移）或迁移说明字符串
    """
    if not old_value:
        return (fallback_module, None)

    v = str(old_value).strip().upper()
    if v in _V11_TO_CURRENT:
        new = _V11_TO_CURRENT[v]
        if new != v:
            return (new, f"v1.1→v1.2: {v} → {new}")
        return (new, None)

    # 不在迁移表里——可能是 v1.2+ 现行枚举或非法值
    return (v, None)


def _assign_ids(cases: list, test_points: dict) -> list:
    """为每个用例分配唯一 ID，按模块分组递增。

    模块定义见 .cursor/MODULES.md（项目级唯一真相源）
    重要：HINT 不作为 Epic ID 前缀，但作为 case_id 前缀允许使用（"HINT-TC-NNN"）
    """
    global _MODULE_COUNTER
    _MODULE_COUNTER = {m: 0 for m in _MODULE_SEQ}

    for case in cases:
        # 中英归一化（"界面" / "UI" / "ui" → "UI"）
        raw_module = case.get("module", "BIZ")
        module = normalize_module_name(raw_module) or "BIZ"

        # 旧枚举迁移（如有）
        if "test_type" in case:
            new_enum, note = migrate_old_enum(case["test_type"], fallback_module=module)
            if note:
                case["_migration_note"] = note
            case["test_type"] = new_enum

        # 旧枚举迁移（"module" 字段本身可能是 v1.1 旧枚举值，常见于 test_points.json）
        if module in _V11_TO_CURRENT and _V11_TO_CURRENT[module] != module:
            new_module = _V11_TO_CURRENT[module]
            case["_migration_note"] = f"v1.1→v1.2: {module} → {new_module}"
            module = new_module

        case["module"] = module  # 写回归一化后的模块名

        seq = _MODULE_COUNTER.get(module, 0) + 1
        _MODULE_COUNTER[module] = seq
        # case_id 前缀 = 归一化后的模块全名
        case["case_id"] = f"{module}-TC-{seq:03d}"

    return cases


def _attach_story_metadata(cases: list, breakdown: dict, test_points: dict) -> list:
    """尽量为用例补齐 story_id / epic_id / module。

    背景：
    - 现有 S6 对话输出不一定显式带 story_id
    - coverage ledger 依赖 story_id 回挂 Story
    - 因此这里做一层基于 backlog / test_points 的最佳努力匹配
    """
    story_catalog: dict[str, dict] = {}
    title_index: dict[str, list[dict]] = {}

    for epic in breakdown.get("epics", []):
        epic_id = epic.get("id", "")
        epic_module = epic.get("module", epic_id.split("-")[0] if epic_id else "BIZ")
        for story in epic.get("stories", []):
            item = {
                "story_id": story.get("id", ""),
                "title": story.get("title", ""),
                "module": epic_module,
                "epic_id": epic_id,
            }
            sid = item["story_id"]
            if sid:
                story_catalog[sid] = item
            norm_title = str(item["title"]).strip().lower()
            if norm_title:
                title_index.setdefault(norm_title, []).append(item)

    if isinstance(test_points, dict):
        for story in test_points.get("stories", []):
            sid = story.get("story_id", "")
            if not sid:
                continue
            story_catalog.setdefault(sid, {
                "story_id": sid,
                "title": story.get("title", ""),
                "module": story.get("module", "BIZ"),
                "epic_id": story.get("epic_id", ""),
            })
            norm_title = str(story.get("title", "")).strip().lower()
            if norm_title:
                title_index.setdefault(norm_title, []).append(story_catalog[sid])

    for case in cases:
        sid = case.get("story_id", "")
        if sid and sid in story_catalog:
            ref = story_catalog[sid]
            case.setdefault("epic_id", ref.get("epic_id", ""))
            case["module"] = normalize_module_name(case.get("module", "")) or normalize_module_name(ref.get("module", "")) or "BIZ"
            continue

        title = (
            case.get("用例描述")
            or case.get("title")
            or case.get("scenario")
            or ""
        )
        norm_title = str(title).strip().lower()
        matched = None
        if norm_title and norm_title in title_index:
            matched = title_index[norm_title][0]
        elif norm_title:
            for candidate_title, items in title_index.items():
                if norm_title in candidate_title or candidate_title in norm_title:
                    matched = items[0]
                    break

        if matched:
            case.setdefault("story_id", matched.get("story_id", ""))
            case.setdefault("epic_id", matched.get("epic_id", ""))
            case["module"] = normalize_module_name(case.get("module", "")) or normalize_module_name(matched.get("module", "")) or "BIZ"
        else:
            case["module"] = normalize_module_name(case.get("module", "")) or "BIZ"

    return cases


def _module_prefix(module: str) -> str:
    """模块 → ID 前缀映射（与 .cursor/MODULES.md §1 严格一致）。

    返回 8 模块英文全名（CONFIG/UI/BIZ/UTIL/LINK/LOG/SPECIAL/HINT）。
    不识别时返回 None（禁止 fallback 到 "MISC"——必须 raise 由 S6 fail_report 处理）。
    """
    return normalize_module_name(module)


def _parse_ai_output(raw: str) -> list:
    """从 AI 输出字符串中解析出用例列表。"""
    cases = []
    # 尝试从 JSON 数组中解析
    json_match = re.search(r'\[[\s\S]+]\]', raw)
    if json_match:
        try:
            cases = json.loads(json_match.group())
            return cases
        except json.JSONDecodeError:
            pass

    # 从 Markdown 表格中解析
    rows = re.findall(r'\|(.+?)\|(.+?)\|(.+?)\|(.+?)\|(.+?)\|', raw)
    for row in rows[1:]:  # 跳过表头
        if len(row) >= 5:
            cases.append({
                "title":        row[0].strip(),
                "module":       row[1].strip(),
                "precondition": row[2].strip(),
                "steps":        row[3].strip(),
                "expected":     row[4].strip(),
            })
    return cases


def _build_summary(cases: list) -> dict:
    """按归一化后的模块名汇总。

    关键修复：原版本用 `case.get("module", "OTHER")`，导致 "界面" / "UI" / "ui"
    三个值在 by_module 中成为 3 个 key。本函数先归一化再分组，确保 8 模块只有 8 个 key。
    """
    by_module = {}
    for c in cases:
        # 归一化：中文/英文缩写/全名 → 8 模块全名
        raw = c.get("module", "BIZ")
        m = normalize_module_name(raw) or "BIZ"
        c["module"] = m  # 反向写回
        by_module.setdefault(m, []).append(c.get("case_id", ""))

    return {
        "total_cases": len(cases),
        "by_module": {m: len(ids) for m, ids in by_module.items()},
        "module_list": sorted(by_module.keys()),  # 排序输出便于对比
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLI: --migrate-modules 框架（历史旧数据平滑过渡到现行枚举）
# ─────────────────────────────────────────────────────────────────────────────

def migrate_test_points_file(json_path: str | Path, *, in_place: bool = False) -> dict:
    """将历史旧 test_points.json / test_cases.json 文件升级到现行 schema。

    升级项：
    1. module 字段：中英归一化（"界面" / "UI" / "ui" → "UI"）
    2. test_point_type 字段：旧枚举 → 现行枚举
    3. 凭空出现的旧枚举（RED_DOT → RED_DOT_BADGE；SYS_MSG → MODAL_DIALOG）

    Args:
        json_path: 待升级的 JSON 文件路径
        in_place: 是否原地写入（默认 False，写到 <原名>.migrated.json）

    Returns:
        dict: {"input": 原文件, "output": 输出文件, "migrations": [...]}
    """
    p = Path(json_path)
    with p.open(encoding="utf-8") as f:
        data = json.load(f)

    migrations = []

    def _migrate_node(node: dict, path: str) -> None:
        # module 归一化
        raw_mod = node.get("module", "")
        if raw_mod:
            new_mod = normalize_module_name(raw_mod)
            if new_mod and new_mod != raw_mod:
                migrations.append({
                    "path": f"{path}.module",
                    "from": raw_mod,
                    "to": new_mod,
                    "reason": "中英归一化",
                })
                node["module"] = new_mod

        # test_point_type / test_type 升级
        for enum_field in ("test_point_type", "test_type", "module_type"):
            raw_enum = node.get(enum_field, "")
            if raw_enum and raw_enum in _V11_TO_CURRENT:
                new_enum = _V11_TO_CURRENT[raw_enum]
                if new_enum != raw_enum:
                    migrations.append({
                        "path": f"{path}.{enum_field}",
                        "from": raw_enum,
                        "to": new_enum,
                        "reason": "v1.1→v1.2/v1.7",
                    })
                    node[enum_field] = new_enum

        # S5 v3 schema 迁移（v1.0 → v3）
        # id → tp_id
        if "id" in node and "tp_id" not in node:
            migrations.append({
                "path": f"{path}.id",
                "from": node["id"],
                "to": node["id"],
                "reason": "v3-schema: id → tp_id（字段重命名）",
            })
            node["tp_id"] = node.pop("id")
        # type → test_point_type
        if "type" in node and "test_point_type" not in node:
            migrations.append({
                "path": f"{path}.type",
                "from": node["type"],
                "to": node["type"],
                "reason": "v3-schema: type → test_point_type（字段重命名）",
            })
            node["test_point_type"] = node.pop("type")
        # title → description（仅对 scenario_test_points 条目生效）
        if "title" in node and "description" not in node:
            migrations.append({
                "path": f"{path}.title",
                "from": node["title"],
                "to": node["title"],
                "reason": "v3-schema: title → description（字段重命名）",
            })
            node["description"] = node.pop("title")

    # 兼容三种结构：
    # 1. test_points.json 顶层 {"stories": [...]}（compose_test_points_from_structure 输出）
    # 2. test_points.json 顶层 [...]（直接列表）
    # 3. test_cases.json 顶层 {"test_cases": [...]}（format_test_cases 输出）

    def _walk_story(story: dict, spath: str) -> None:
        _migrate_node(story, spath)
        for tp_idx, tp in enumerate(story.get("scenario_test_points", [])):
            _migrate_node(tp, f"{spath}.scenario_test_points[{tp_idx}]")
        # module_coverage 内部 points[] 也要迁移
        mc = story.get("module_coverage", {})
        if isinstance(mc, dict):
            for mod_key, mc_val in mc.items():
                if not isinstance(mc_val, dict):
                    continue
                new_mod = normalize_module_name(mod_key)
                if new_mod and new_mod != mod_key:
                    migrations.append({
                        "path": f"{spath}.module_coverage.{mod_key}",
                        "from": mod_key,
                        "to": new_mod,
                        "reason": "中英归一化",
                    })
                    mc[new_mod] = mc_val
                    if new_mod != mod_key:
                        del mc[mod_key]
                for p_idx, p in enumerate(mc_val.get("points", [])):
                    if p in _V11_TO_CURRENT:
                        new_p = _V11_TO_CURRENT[p]
                        if new_p != p:
                            migrations.append({
                                "path": f"{spath}.module_coverage.{new_mod or mod_key}.points[{p_idx}]",
                                "from": p,
                                "to": new_p,
                                "reason": "v1.1→v1.2/v1.7",
                            })
                            mc_val["points"][p_idx] = new_p

    if isinstance(data, list):
        # 顶层列表
        for s_idx, story in enumerate(data):
            _walk_story(story, f"[{s_idx}]")
    elif isinstance(data, dict):
        if "stories" in data:
            for s_idx, story in enumerate(data["stories"]):
                _walk_story(story, f"stories[{s_idx}]")
        if "test_cases" in data:
            for tc_idx, tc in enumerate(data["test_cases"]):
                _migrate_node(tc, f"test_cases[{tc_idx}]")

    # 写出
    out_path = p if in_place else p.with_suffix(".migrated.json")
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return {
        "input":  str(p),
        "output": str(out_path),
        "migrations_count": len(migrations),
        "migrations": migrations,
    }


def _cli_migrate_modules(argv: list[str]) -> int:
    """CLI 入口：
    `python test_case_formatter.py --migrate-modules <json> [--in-place]`
    迁移内容：
    1. v1.1→v1.2/v1.7 旧枚举升级（module / test_point_type）
    2. S5 v3 schema（v1.0→v3）：id→tp_id, type→test_point_type, title→description
    """
    import argparse
    ap = argparse.ArgumentParser(
        prog="test_case_formatter.py",
        description="AIDocxWorkFlow test_case_formatter CLI — v1.1 → v1.2/v1.7 迁移",
    )
    ap.add_argument("--migrate-modules", metavar="JSON", help="升级 v1.1 旧 JSON 文件")
    ap.add_argument("--in-place", action="store_true", help="原地写入（默认输出 <原名>.migrated.json）")
    ap.add_argument("--tc-json-to-xlsx", metavar="JSON", help="按公共 10 列格式导出双 Sheet xlsx")
    ap.add_argument("--xlsx-output", metavar="XLSX", help="指定 xlsx 输出路径（不可覆盖表头）")
    args = ap.parse_args(argv)

    if args.tc_json_to_xlsx:
        output = export_test_cases_json_to_xlsx(args.tc_json_to_xlsx, args.xlsx_output)
        print(f"[tc-json-to-xlsx] input  : {args.tc_json_to_xlsx}")
        print(f"[tc-json-to-xlsx] output : {output}")
        return 0

    if args.migrate_modules:
        result = migrate_test_points_file(args.migrate_modules, in_place=args.in_place)
        print(f"[migrate-modules] input  : {result['input']}")
        print(f"[migrate-modules] output : {result['output']}")
        print(f"[migrate-modules] migrations: {result['migrations_count']} 项")
        for m in result["migrations"][:20]:
            print(f"  - {m['path']}: {m['from']} → {m['to']} ({m['reason']})")
        if len(result["migrations"]) > 20:
            print(f"  ... 还有 {len(result['migrations']) - 20} 项")
        return 0

    ap.print_help()
    return 1


def _save_md(cases: list, output_dir: Path) -> Path:
    """保存 10 列 Markdown：与 SKILL.md 当前 10 列严格一致。

    重要：现行 10 列版本
    - 旧 6 列（ID/标题/模块/前置/步骤/预期）→ 10 列
    - 缺「功能描述」「用例状态」「备注」会导致 S7 审查丢分
    """
    lines = [
        "# 测试用例 — 游戏道具商城系统 v1.0\n",
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n",
        f"用例总数：{len(cases)}\n\n",
        "## 用例列表\n\n",
        "| 用例ID | 模块 | 用例描述 | 功能描述 | 前置条件 | 操作步骤 | 预期结果 | 优先级 | 用例状态 | 备注 |\n",
        "|---|---|---|---|---|---|---|---|---|---|\n",
    ]
    for c in cases:
        lines.append(
            f"| {_get_field(c, 'case_id')} | {_get_field(c, '模块', 'module')} | "
            f"{_get_field(c, '用例描述', 'scenario')} | {_get_field(c, '功能描述', 'description')} | "
            f"{_get_field(c, '前置条件', 'precondition')} | "
            f"{_get_field(c, '操作步骤', 'test_steps', 'steps')} | "
            f"{_get_field(c, '预期结果', 'expected_result', 'expected')} | "
            f"{_get_field(c, '优先级', 'priority', _default='P2')} | "
            f"{_get_field(c, '用例状态', 'case_status')} | "
            f"{_get_field(c, '备注', 'story_id')} |\n"
        )

    md_path = output_dir / "test_cases.md"
    with md_path.open("w", encoding="utf-8") as f:
        f.writelines(lines)
    return md_path


_XLSX_HEADERS_V3 = [
    "用例ID",     # case_id
    "模块",       # 模块 / module
    "用例描述",   # 用例描述（纯 Story 标题）
    "功能描述",   # 功能描述（scenario）
    "前置条件",   # 前置条件 / precondition
    "操作步骤",   # 操作步骤（test_steps / steps）
    "预期结果",   # 预期结果（expected_result / expected）
    "优先级",     # 优先级 P0/P1/P2
    "用例状态",   # 用例状态 Draft/Ready/Rejected/Deprecated
    "备注",       # 备注 story_id/tp_id/method 元数据
]

_DEFAULT_XLSX_PROFILE = {
    "sheet_name": "测试用例",
    "headers": _XLSX_HEADERS_V3,
    "field_mapping": {
        "用例ID": ["case_id"],
        "模块": ["模块", "module"],
        "用例描述": ["用例描述", "scenario"],
        "功能描述": ["功能描述", "description"],
        "前置条件": ["前置条件", "preconditions", "precondition"],
        "操作步骤": ["操作步骤", "test_steps", "steps"],
        "预期结果": ["预期结果", "expected_results", "expected_result", "expected"],
        "优先级": ["优先级", "priority"],
        "用例状态": ["用例状态", "case_status"],
        "备注": ["备注", "story_id"],
    },
}

_ROUND17_XLSX_HEADERS = [
    "用例ID",
    "所属模块",
    "用例描述",
    "关联需求",
    "功能描述",
    "前置条件",
    "操作步骤",
    "预期结果",
    "优先级",
    "用例状态",
    "备注",
]

_ROUND17_XLSX_PROFILE = copy.deepcopy(_DEFAULT_XLSX_PROFILE)
_ROUND17_XLSX_PROFILE["headers"] = _ROUND17_XLSX_HEADERS
_ROUND17_XLSX_PROFILE["field_mapping"].update({
    "所属模块": ["模块", "module"],
    "关联需求": ["obj_id"],
})


def _project_s6_root(project_name: str | None) -> Path | None:
    if not project_name:
        return None
    return PROJECT_KNOWLEDGE_ROOT / project_name / "s6"


def _load_project_export_profile(project_name: str | None) -> dict[str, Any] | None:
    """读取项目级导出配置。未确认项目时，默认只允许输出公共 JSON。"""
    project_root = _project_s6_root(project_name)
    if project_root is None:
        return None
    profile_path = project_root / "export_profiles" / "test_cases.export.json"
    if not profile_path.exists():
        return None
    try:
        profile = json.loads(profile_path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(profile, dict):
        return None
    merged = copy.deepcopy(_DEFAULT_XLSX_PROFILE)
    merged.update({k: v for k, v in profile.items() if k != "field_mapping"})
    if isinstance(profile.get("field_mapping"), dict):
        merged["field_mapping"].update(profile["field_mapping"])
    return merged


def _load_project_xlsx_template(project_name: str | None) -> Path | None:
    project_root = _project_s6_root(project_name)
    if project_root is None:
        return None
    template_path = project_root / "xlsx_templates" / "test_cases.template.xlsx"
    return template_path if template_path.exists() else None


def _get_field(case: dict, *keys, _default: str = "") -> str:
    """从 case dict 中按优先级查找第一个存在的字段值。

    兼容两套字段名体系：
    - SKILL.md 中文体系：用例描述 / 功能描述 / 操作步骤 / 预期结果
    - S6 对话直接生成体系：scenario / expected_result / test_steps / story_id
    - 历史 S6 体系：steps / expected

    无匹配时返回空字符串；外部可通过 default= 参数覆盖。

    list 元素的渲染（v3.01 修复）：
    - str 元素直接拼
    - dict 元素（如 ``{"step_num": 1, "action": "玩家点击商城入口"}``）
      格式化为 ``"{step_num}. {action}"``——避免 Excel 单元格出现 Python dict repr
    - 其它类型回退到 ``str(item)``

    Dict-repr 字符串 fall-through（v3.01 修复 · Round 2 Act）：
    若字段值是 str 且看起来是 dict repr（以 ``{`` 开头并以 ``}`` 结尾），
    说明 mirror_bilingual_aliases 已把 list 错误地 join 成字符串——
    此时应 fall through 到下一个 key（避免返回 ``"{'step_num': 1, ...}"`` 字符串）。
    """
    for key in keys:
        val = case.get(key)
        if val is None:
            continue
        # dict-repr 字符串 → fall through（list 字段被错误 mirror 成字符串）
        if isinstance(val, str) and val.startswith("{") and val.endswith("}"):
            continue
        # test_steps / steps 是 list，转 "\n" 分行字符串
        if isinstance(val, list):
            return "\n".join(_render_list_item(item, index) for index, item in enumerate(val))
        return str(val)
    return _default


def _render_list_item(item: Any, index: int) -> str:
    """Render one element of a list-valued TC field (操作步骤 / 前置条件 / 预期结果).

    v3.01 修复：dict 元素（step 字典）格式化为 ``"{step_num}. {action}"``，而不是
    直接 ``str(item)``——后者会输出 ``"{'step_num': 1, 'action': '...'}"`` 的
    Python dict repr 到 Excel 单元格，导致 xlsx 不可读（Q-007 BLOCKER）。

    v2.0 扩展：支持新格式预期字典 ``{step_ref, 预期}``，渲染为 ``"[步骤{sn}] {预期}"``。

    对于 dict 元素，识别以下常见键名（按 SSOT STAGE_S6_TEST_CASES.mdc §11）：
    - 操作步骤格式：``step_num`` / ``step_no`` / ``index`` / ``seq`` → 步骤号
      + ``action`` / ``description`` / ``step`` / ``cmd`` → 步骤动作
    - 预期结果格式：``step_ref`` → 关联步骤号
      + ``预期`` / ``expected`` → 预期文本
    """
    if isinstance(item, dict):
        # 检测是否为新格式预期字典（含 step_ref 和 预期/expected 字段）
        step_ref = item.get("step_ref")
        expected_text = item.get("预期") or item.get("expected")
        if step_ref is not None and expected_text:
            # 新格式预期：渲染为 "[步骤{sn}] {预期}"
            return f"[步骤{step_ref}] {expected_text}"

        # 操作步骤格式
        step_num = (
            item.get("step_num")
            or item.get("step_no")
            or item.get("index")
            or item.get("seq")
            or (index + 1)
        )
        action = (
            item.get("action")
            or item.get("description")
            or item.get("step")
            or item.get("cmd")
            or ""
        )
        try:
            return f"{step_num}. {action}" if action else f"{step_num}."
        except Exception:
            return str(item)
    return str(item)


# ─────────────────────────────────────────────────────────────────────────────
# v28: TC 内部结构化映射 — 步骤-预期对应渲染
# ─────────────────────────────────────────────────────────────────────────────


def is_new_structural_mapping_format(tc: dict) -> bool:
    """检测 TC 是否为新的层级结构格式（v28 TC 结构化映射规范）。

    新格式特征：
    - 预期结果含 step_ref 字段（dict 数组）
    - 步骤数 ≥ 3（禁止 1 步 1 TC）

    Args:
        tc: 单条 TC dict

    Returns:
        True: 新格式（有 step_ref 或步骤数符合要求）
    """
    # 检测预期结果是否含 step_ref
    expected = tc.get("预期结果") or tc.get("expected_results") or []
    if expected and isinstance(expected, list) and len(expected) > 0:
        first = expected[0]
        if isinstance(first, dict) and "step_ref" in first:
            return True

    # 检测步骤数是否 ≥ 3（间接判断）
    steps = tc.get("操作步骤") or tc.get("test_steps") or tc.get("steps") or []
    if isinstance(steps, list) and len(steps) >= 3:
        return True

    return False


def render_steps_with_expected(steps: list, expected: list) -> tuple[str, str]:
    """渲染步骤与预期，返回 (steps_text, expected_text)。

    v28 TC 结构化映射规范：
    - 步骤渲染：按 step_num 顺序输出
    - 预期渲染：按 step_ref 对应步骤号输出

    Args:
        steps: 操作步骤列表 [{step_num, action}, ...]
        expected: 预期结果列表 [{step_ref, 预期}, ...] 或 [文本, ...]

    Returns:
        (steps_text, expected_text) - 两列的多行文本
    """
    # 步骤渲染
    step_lines = []
    for idx, s in enumerate(steps):
        if isinstance(s, dict):
            sn = s.get("step_num") or s.get("step_no") or (idx + 1)
            act = s.get("action") or s.get("description") or s.get("step") or ""
            step_lines.append(f"{sn}. {act}")
        else:
            step_lines.append(str(s))

    # 预期渲染
    if expected and isinstance(expected, list) and len(expected) > 0:
        first = expected[0]
        if isinstance(first, dict) and "step_ref" in first:
            # 新格式：按 step_ref 对应输出
            exp_lines = []
            for e in expected:
                sn = e.get("step_ref", "")
                txt = e.get("预期") or e.get("expected") or ""
                exp_lines.append(f"[步骤{sn}] {txt}")
        else:
            # 旧格式：直接输出文本
            exp_lines = [str(e) for e in expected]
    else:
        exp_lines = []

    return "\n".join(step_lines), "\n".join(exp_lines)


def render_new_format_tc(case: dict) -> dict:
    """将新格式 TC（简化版）转换为 Excel 渲染格式。

    处理步骤-预期对应渲染，返回可用于 _xlsx_row 的字段。

    Args:
        case: 原始 TC dict

    Returns:
        渲染后的 dict，含 "操作步骤" 和 "预期结果" 字符串
    """
    steps = case.get("操作步骤") or case.get("test_steps") or case.get("steps") or []
    expected = case.get("预期结果") or case.get("expected_results") or []

    steps_text, expected_text = render_steps_with_expected(steps, expected)

    # 返回渲染后的 dict（不修改原始）
    result = dict(case)
    result["操作步骤"] = steps_text
    result["预期结果"] = expected_text
    return result


def validate_structural_mapping(tc: dict) -> dict:
    """校验 TC 的结构化映射合规性。

    v28 TC 结构化映射规范自检清单：
    1. 步骤数 ≥ 2（禁止 1 步 1 TC）
    2. 预期数 ≥ 1
    3. 预期与步骤有对应（显式 step_ref 或顺序）

    Args:
        tc: 单条 TC dict

    Returns:
        {"passed": bool, "errors": [错误列表]}
    """
    errors = []

    steps = tc.get("操作步骤") or tc.get("test_steps") or tc.get("steps") or []
    expected = tc.get("预期结果") or tc.get("expected_results") or []

    # 规则 1: 步骤数 ≥ 2
    if isinstance(steps, list) and len(steps) < 2:
        errors.append(f"步骤数不足：{len(steps)} < 2（禁止 1 步 1 TC）")

    # 规则 2: 预期数 ≥ 1
    if isinstance(expected, list) and len(expected) < 1:
        errors.append("预期结果为空")

    # 规则 3: 预期与步骤对应
    if expected and isinstance(expected, list) and len(expected) > 0:
        first = expected[0]
        if isinstance(first, dict) and "step_ref" in first:
            # 有 step_ref，检查是否所有 step_ref 都在步骤范围内
            step_nums = {s.get("step_num") or idx + 1 for idx, s in enumerate(steps) if isinstance(s, dict)}
            for e in expected:
                ref = e.get("step_ref")
                if ref and ref not in step_nums:
                    errors.append(f"step_ref={ref} 超出步骤范围 {step_nums}")

    return {"passed": len(errors) == 0, "errors": errors}


def _partition_cases_for_xlsx(cases: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Split cases into the Ready main sheet and Draft/Rejected appendix.

    Note
    ----
    支持 legacy English 字段别名 — 调用方需先经
    ``ai_workflow.case_id_and_field_normalizer.normalize_payload`` 归一化，
    再调 ``_save_xlsx``；本函数本身不读 ``preconditions`` / ``steps`` 等英文
    字段（这些是 normalizer 已镜像到中文 ``前置条件`` / ``操作步骤`` 后再传入
    的）。属于「下游消费方」而非「归一化方」。
    """
    ready = [case for case in cases if _get_field(case, "用例状态", "case_status") == "Ready"]
    appendix = [
        case for case in cases
        if _get_field(case, "用例状态", "case_status") in {"Draft", "Rejected"}
    ]
    return ready, appendix


def _xlsx_row(
    case: dict[str, Any],
    headers: list[str],
    field_mapping: dict[str, list[str]],
) -> list[str]:
    row: list[str] = []
    for header in headers:
        keys = field_mapping.get(header, [])
        if header == "优先级":
            row.append(_get_field(case, *keys, _default="P2"))
        else:
            row.append(_get_field(case, *keys))
    return row


def _populate_worksheet(
    worksheet: Any,
    headers: list[str],
    cases: list[dict[str, Any]],
    field_mapping: dict[str, list[str]],
) -> None:
    worksheet.append(headers)
    for case in cases:
        worksheet.append(_xlsx_row(case, headers, field_mapping))
    for column in worksheet.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)


# ---------------------------------------------------------------------------
# Round 17 — OBJ grouping: sorted rows, color bands, vertical merges, no spacers
# ---------------------------------------------------------------------------

# Alternating fills for consecutive OBJ groups (light-gray / light-blue /
# light-yellow / light-green). Chosen to be readable in both Excel and
# LibreOffice and to pass WCAG AA contrast on default text.
_OBJ_GROUP_FILLS: tuple[str, ...] = (
    "FFE6E6E6",  # light gray
    "FFE3F2FD",  # light blue
    "FFFFF8E1",  # light yellow
    "FFE8F5E9",  # light green
    "FFF3E5F5",  # light purple
)


def _sort_cases_by_obj_fp(
    cases: list[dict[str, Any]],
    sort_options: dict[str, Any],
) -> list[dict[str, Any]]:
    """Stable-sort cases by ``obj_id`` → ``feature_point_ref`` → ``case_id``.

    Empty / missing sort keys sort to the end (defensive — legacy data may
    omit ``obj_id`` or ``feature_point_ref``). The function is purely
    functional — it returns a new list, leaving the input untouched.
    """
    sort_keys = sort_options.get("by") or ["obj_id", "feature_point_ref", "case_id"]

    def _sort_tuple(case: Mapping[str, Any]) -> tuple[str, ...]:
        vals: list[str] = []
        for key in sort_keys:
            v = case.get(key)
            if v is None:
                vals.append("\uffff")  # sort to end
            else:
                vals.append(str(v))
        return tuple(vals)

    return sorted(cases, key=_sort_tuple)


def _block_is_field_uniform(
    block_cases: list[dict[str, Any]],
    header: str,
    field_mapping: dict[str, list[str]],
    headers: list[str],
) -> bool:
    """Return True iff every case in ``block_cases`` has the same value for
    the column identified by ``header`` (resolved via ``field_mapping``).

    Used by Round 18 to guard against destructive merges when the OBJ block's
    column is per-TC/FP rather than per-OBJ. Reads from the in-memory case
    list (NOT the worksheet) so that merged-cell ``None`` side-effects cannot
    corrupt the check.

    Empty / missing values count as uniform-with-each-other only when EVERY
    row is empty — otherwise we treat the column as non-uniform to be safe.
    """
    if not block_cases:
        return False
    keys = field_mapping.get(header, [])
    if not keys:
        return False
    seen: set[str] = set()
    for case in block_cases:
        # Resolve value via the same _get_field fallback chain _xlsx_row uses,
        # so we compare exactly what would be written to the cell.
        if header == "优先级":
            val = _get_field(case, *keys, _default="P2")
        else:
            val = _get_field(case, *keys)
        if val in (None, ""):
            continue
        seen.add(str(val))
        if len(seen) > 1:
            return False
    return len(seen) == 1


def _populate_worksheet_with_obj_grouping(
    worksheet: Any,
    headers: list[str],
    cases: list[dict[str, Any]],
    field_mapping: dict[str, list[str]],
    sort_options: dict[str, Any],
) -> None:
    """Populate the main sheet with contiguous, color-banded OBJ blocks.

    ``cases`` must already be sorted by :func:`_sort_cases_by_obj_fp`. Rows are
    written without spacers. When ``merge_obj_headers`` is supplied, each
    multi-row OBJ block is vertically merged for those headers after all rows
    have been written. Single-row OBJ blocks keep ordinary cells.
    """
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, PatternFill  # type: ignore

    obj_color_band = sort_options.get("obj_color_band", True)
    merge_obj_headers = [
        str(header)
        for header in sort_options.get("merge_obj_headers", [])
        if str(header) in headers
    ]

    worksheet.append(headers)
    if hasattr(worksheet, "cell"):
        for col_idx in range(1, len(headers) + 1):
            try:
                worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
            except Exception:  # pragma: no cover
                pass

    prev_obj_id: str | None = None
    obj_color_index = -1
    current_fill = None
    group_start_row: int | None = None
    obj_blocks: list[tuple[int, int, Any, list[dict[str, Any]]]] = []
    _block_cases: list[dict[str, Any]] = []

    for case in cases:
        obj_id = str(case.get("obj_id") or "")
        if obj_id != prev_obj_id:
            if prev_obj_id is not None and group_start_row is not None:
                obj_blocks.append((group_start_row, worksheet.max_row, current_fill, _block_cases))
            obj_color_index = (obj_color_index + 1) % len(_OBJ_GROUP_FILLS)
            prev_obj_id = obj_id
            group_start_row = worksheet.max_row + 1
            current_fill = (
                PatternFill(
                    start_color=_OBJ_GROUP_FILLS[obj_color_index],
                    end_color=_OBJ_GROUP_FILLS[obj_color_index],
                    fill_type="solid",
                )
                if obj_color_band
                else None
            )
            _block_cases = []
        _block_cases.append(case)

        worksheet.append(_xlsx_row(case, headers, field_mapping))
        if current_fill is not None and hasattr(worksheet, "cell"):
            row_idx = worksheet.max_row
            for col_idx in range(1, len(headers) + 1):
                try:
                    worksheet.cell(row=row_idx, column=col_idx).fill = current_fill
                except Exception:  # pragma: no cover
                    pass

    if prev_obj_id is not None and group_start_row is not None:
        obj_blocks.append((group_start_row, worksheet.max_row, current_fill, _block_cases))

    header_columns = {header: headers.index(header) + 1 for header in merge_obj_headers}
    for start_row, end_row, block_fill, block_cases in obj_blocks:
        if end_row > start_row:
            # Round 18: merge ONLY columns whose values are uniform across the
            # entire OBJ block. The user's report (R18-B1) requested merging
            # 5 columns (用例描述 / 所属模块 / 关联需求 / 功能描述 / 前置条件),
            # but physical data shows E (功能描述) and F (前置条件) are FP-level,
            # not OBJ-level — 16/16 OBJ blocks have non-uniform E and F. A naive
            # merge would destroy data. We honour the user's request by extending
            # merge_obj_headers, but apply a data-preservation filter that only
            # merges when values match across all rows in the block.
            for header, col_idx in header_columns.items():
                if _block_is_field_uniform(block_cases, header, field_mapping, headers):
                    worksheet.merge_cells(
                        start_row=start_row,
                        start_column=col_idx,
                        end_row=end_row,
                        end_column=col_idx,
                    )
                    worksheet.cell(row=start_row, column=col_idx).alignment = Alignment(
                        vertical="center",
                        wrap_text=True,
                    )
        if block_fill is not None:
            for row_idx in range(start_row, end_row + 1):
                for col_idx in range(1, len(headers) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = block_fill

    for column in worksheet.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)


def _build_case_description_index_rows(
    cases: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Aggregate Ready cases by ``obj_id`` → ``用例描述`` for the description-index Sheet 2.

    Returns a list of dicts with keys:
    - ``obj_id``
    - ``obj_name`` (用例描述, OBJ 名称)
    - ``fp_count`` — number of distinct ``feature_point_ref`` values
    - ``tc_count`` — total TCs (all statuses; passed in)
    - ``ready_count`` — number of TCs with ``用例状态 == "Ready"``

    Sorted by ``obj_id`` ascending, then by ``obj_name`` ascending so the
    index sheet reads top-to-bottom in the same grouping order as the main
    sheet.
    """
    aggregates: dict[tuple[str, str], dict[str, Any]] = {}
    for case in cases:
        if not isinstance(case, Mapping):
            continue
        obj_id = str(case.get("obj_id") or "")
        obj_name = str(case.get("用例描述") or "")
        fp_ref = str(case.get("feature_point_ref") or "")
        is_ready = str(case.get("用例状态") or case.get("case_status") or "") == "Ready"
        key = (obj_id, obj_name)
        bucket = aggregates.setdefault(key, {
            "obj_id": obj_id,
            "obj_name": obj_name,
            "_fps": set(),
            "_tc_count": 0,
            "_ready_count": 0,
        })
        if fp_ref:
            bucket["_fps"].add(fp_ref)
        bucket["_tc_count"] += 1
        if is_ready:
            bucket["_ready_count"] += 1
    rows: list[dict[str, Any]] = []
    for bucket in aggregates.values():
        rows.append({
            "obj_id": bucket["obj_id"],
            "obj_name": bucket["obj_name"],
            "fp_count": len(bucket["_fps"]),
            "tc_count": bucket["_tc_count"],
            "ready_count": bucket["_ready_count"],
        })
    rows.sort(key=lambda r: (r["obj_id"], r["obj_name"]))
    return rows


_INDEX_SHEET_HEADERS: tuple[str, ...] = (
    "OBJ ID",
    "用例描述(OBJ 名)",
    "FP 数",
    "TC 数",
    "Ready 数",
)


def _populate_description_index_sheet(
    worksheet: Any,
    rows: list[dict[str, Any]],
) -> None:
    """Write the description-index Sheet 2.

    Columns: OBJ ID / 用例描述 (OBJ 名) / FP 数 / TC 数 / Ready 数. Header is
    bold; rows are written in input order (already pre-sorted by the caller).
    """
    import openpyxl  # type: ignore

    worksheet.append(list(_INDEX_SHEET_HEADERS))
    for row_idx, row in enumerate(rows, start=2):
        worksheet.append([
            row["obj_id"],
            row["obj_name"],
            row["fp_count"],
            row["tc_count"],
            row["ready_count"],
        ])
    if hasattr(worksheet, "cell"):
        for col_idx in range(1, len(_INDEX_SHEET_HEADERS) + 1):
            try:
                worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
            except Exception:  # pragma: no cover
                pass
    # Column widths tuned for the index content.
    widths = (16, 36, 10, 10, 10)
    for column in worksheet.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=0)
        col_letter = column[0].column_letter
        # Take the larger of measured width and the per-column tuned width.
        tuned = widths[ord(col_letter) - ord("A")] if len(column[0].column_letter) == 1 else 16
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, tuned), 50)


def _save_xlsx(
    cases: list,
    output_dir: Path,
    *,
    export_profile: dict[str, Any] | None = None,
    template_path: Path | None = None,
    output_path: Path | str | None = None,
    sort_options: dict[str, Any] | None = None,
    description_index_sheet: bool = False,
) -> Path:
    """Save Ready cases to the main sheet and Draft/Rejected to an appendix.

    Round 17 extensions remain opt-in; default behavior is unchanged:

    - ``sort_options`` (dict, optional): when provided, the main-sheet Ready
      cases are sorted by the configured keys and receive contiguous OBJ color
      bands without spacer rows. Keys:

      - ``"by"``: list of field names to sort by (default
        ``["obj_id", "feature_point_ref", "case_id"]``)
      - ``"obj_color_band"``: bool, default ``True``
      - ``"merge_obj_headers"``: header names to merge vertically for every
        multi-row OBJ block

    - ``description_index_sheet`` (bool, default ``False``): when True, an
      additional ``用例描述索引`` sheet is appended. It aggregates Ready cases
      by ``obj_id`` and OBJ name, then shows FP / TC / Ready counts.

    Both flags remain off by default. The Ready-main / Draft-Rejected-appendix
    partition contract is unchanged; sorting, color bands, and OBJ merges apply
    only to the main sheet.
    """
    if isinstance(output_dir, str):
        output_dir = Path(output_dir)
    xlsx_path = Path(output_path) if output_path else output_dir / "test_cases.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    profile = export_profile or _DEFAULT_XLSX_PROFILE
    headers = profile.get("headers", _XLSX_HEADERS_V3)
    field_mapping = profile.get("field_mapping", _DEFAULT_XLSX_PROFILE["field_mapping"])
    main_sheet_name = profile.get("sheet_name", "测试用例")
    appendix_sheet_name = "Draft-Rejected附录"
    ready_cases, appendix_cases = _partition_cases_for_xlsx(cases)
    sheets = [
        (main_sheet_name, ready_cases),
        (appendix_sheet_name, appendix_cases),
    ]

    # Opt-in main-sheet grouping: sort Ready cases and apply contiguous OBJ
    # color bands plus any requested shared-header merges. Default
    # (sort_options=None) preserves legacy input order and layout.
    main_grouping_enabled = sort_options is not None
    if main_grouping_enabled:
        ready_cases = _sort_cases_by_obj_fp(ready_cases, sort_options or {})

    # Round 16 opt-in: build a separate "case-description index" Sheet 2 that
    # aggregates Ready cases by OBJ and exposes FP / TC / Ready counts.
    index_rows: list[dict[str, Any]] = []
    if description_index_sheet:
        index_rows = _build_case_description_index_rows(ready_cases)

    try:
        import openpyxl

        if template_path and template_path.exists():
            workbook = openpyxl.load_workbook(str(template_path))
            main_sheet = (
                workbook[main_sheet_name]
                if main_sheet_name in workbook.sheetnames
                else workbook[workbook.sheetnames[0]]
            )
            main_sheet.title = main_sheet_name
            main_sheet.delete_rows(1, main_sheet.max_row)
            if appendix_sheet_name in workbook.sheetnames:
                del workbook[appendix_sheet_name]
        else:
            workbook = openpyxl.Workbook()
            main_sheet = workbook.active
            main_sheet.title = main_sheet_name

        if main_grouping_enabled:
            _populate_worksheet_with_obj_grouping(
                main_sheet, headers, ready_cases, field_mapping, sort_options or {}
            )
        else:
            _populate_worksheet(main_sheet, headers, ready_cases, field_mapping)
        appendix_sheet = workbook.create_sheet(appendix_sheet_name)
        _populate_worksheet(appendix_sheet, headers, appendix_cases, field_mapping)
        # Round 16: append the description-index Sheet 2 after the main sheet
        # and appendix so it lands at index 2 (third sheet). We position it
        # explicitly so users opening the file see: 测试用例 / Draft-Rejected附录
        # / 用例描述索引. When appendix is empty, the order collapses to
        # 测试用例 / 用例描述索引 — also fine for end users.
        if description_index_sheet and index_rows:
            index_sheet = workbook.create_sheet("用例描述索引")
            _populate_description_index_sheet(index_sheet, index_rows)
        workbook.save(str(xlsx_path))
    except ImportError:
        rows_by_sheet = [
            (sheet_name, [headers] + [_xlsx_row(case, headers, field_mapping) for case in sheet_cases])
            for sheet_name, sheet_cases in sheets
        ]
        _build_xlsx_archive(xlsx_path, rows_by_sheet)
    return xlsx_path


def _load_test_cases_payload(json_path: Path | str) -> list[dict[str, Any]]:
    """Load one of the three supported public test-case JSON schemas."""
    path = Path(json_path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        cases = payload
    elif isinstance(payload, dict) and isinstance(payload.get("test_cases"), list):
        cases = payload["test_cases"]
    else:
        raise ValueError("tc.json must be a list, {test_cases:[...]}, or {meta,test_cases}")
    if not all(isinstance(case, dict) for case in cases):
        raise ValueError("every test_cases item must be an object")
    return cases


def export_test_cases_json_to_xlsx(
    json_path: Path | str,
    output_path: Path | str | None = None,
) -> Path:
    """Export tc.json with the immutable public profile and two-sheet layout."""
    source = Path(json_path)
    destination = Path(output_path) if output_path else source.with_suffix(".xlsx")
    cases = _load_test_cases_payload(source)
    return _save_xlsx(
        cases,
        destination.parent,
        export_profile=copy.deepcopy(_DEFAULT_XLSX_PROFILE),
        template_path=None,
        output_path=destination,
    )


def _write_minimal_xlsx(
    xlsx_path: Path,
    headers: list[str],
    cases: list[dict[str, Any]],
    field_mapping: dict[str, list[str]],
    sheet_name: str,
) -> None:
    ready_cases, appendix_cases = _partition_cases_for_xlsx(cases)
    rows_by_sheet = [
        (sheet_name, [headers] + [_xlsx_row(case, headers, field_mapping) for case in ready_cases]),
        (
            "Draft-Rejected附录",
            [headers] + [_xlsx_row(case, headers, field_mapping) for case in appendix_cases],
        ),
    ]
    _build_xlsx_archive(xlsx_path, rows_by_sheet)


def _xlsx_column_name(index: int) -> str:
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _build_xlsx_archive(
    xlsx_path: Path,
    rows_by_sheet: list[tuple[str, list[list[str]]]],
) -> None:
    """Build a minimal multi-sheet xlsx archive when openpyxl is unavailable."""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    worksheet_xml: list[str] = []
    sheet_nodes: list[str] = []
    worksheet_relationships: list[str] = []
    content_overrides: list[str] = []

    for sheet_index, (sheet_name, rows) in enumerate(rows_by_sheet, start=1):
        sheet_rows = []
        for row_index, row in enumerate(rows, start=1):
            cells = []
            for column_index, value in enumerate(row, start=1):
                ref = f"{_xlsx_column_name(column_index)}{row_index}"
                cells.append(
                    f'<c r="{ref}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>'
                )
            sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
        width = len(rows[0]) if rows else 1
        height = len(rows) if rows else 1
        worksheet_xml.append(
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<dimension ref="A1:{_xlsx_column_name(width)}{height}"/>'
            '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
            '<sheetFormatPr defaultRowHeight="15"/>'
            f'<sheetData>{"".join(sheet_rows)}</sheetData>'
            '</worksheet>'
        )
        sheet_nodes.append(
            f'<sheet name="{escape(sheet_name)}" sheetId="{sheet_index}" r:id="rId{sheet_index}"/>'
        )
        worksheet_relationships.append(
            f'<Relationship Id="rId{sheet_index}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{sheet_index}.xml"/>'
        )
        content_overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{sheet_index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )

    style_rel_id = len(rows_by_sheet) + 1
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets>{"".join(sheet_nodes)}</sheets>'
        '</workbook>'
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(worksheet_relationships)}'
        f'<Relationship Id="rId{style_rel_id}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
        '</Relationships>'
    )
    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        '</Relationships>'
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        '<borders count="1"><border/></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        f'{"".join(content_overrides)}'
        '<Override PartName="/xl/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '</Types>'
    )
    with zipfile.ZipFile(xlsx_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", root_rels_xml)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        for sheet_index, sheet_xml in enumerate(worksheet_xml, start=1):
            archive.writestr(f"xl/worksheets/sheet{sheet_index}.xml", sheet_xml)
        archive.writestr("xl/styles.xml", styles_xml)


# ─────────────────────────────────────────────────────────────────────────────
# v10: S2→S6 OBJ 链接强制条款的 L3 兜底校验
# ─────────────────────────────────────────────────────────────────────────────


def _load_s2_requirement_objects(req_name: str, version: str) -> list[dict] | None:
    """加载 S2 需求对象列表——v10 OBJ 链接校验的输入源。

    返回 None 表示无法加载（目录不存在 / 文件缺失 / JSON 解析失败）。
    """
    assets_root = _ROOT / "workflow_assets" / req_name / version / "「S2 需求拆解」"
    obj_path = assets_root / "requirement_objects.json"
    if not obj_path.exists():
        return None
    try:
        data = json.loads(obj_path.read_text(encoding="utf-8"))
        objs = data.get("requirement_objects", [])
        return objs if isinstance(objs, list) else None
    except Exception:
        return None


def _validate_obj_linkage(
    test_cases: list[dict],
    s2_objs: list[dict],
) -> dict[str, Any]:
    """S6 TC 与 S2 OBJ 链接校验（v10 §OBJ 链接字段定义 L3 兜底）。

    门禁（违反任一 → passed=False）：
    - 每个 TC 有 obj_id（缺失 → missing_obj_id 列出 case_id）
    - 每个 obj_id 在 S2 OBJ 列表中存在（不存在的 → invalid_obj_id）
    - 每个 TC 的用例描述 == obj_name（不等 → desc_mismatch 列出）
    - obj_linkage_coverage = TC 引用的 OBJ 数 / S2 OBJ 总数 ≥ 1.0（未达 → coverage < 1.0）

    返回结构供 postflight_gate.json 直接写入 obj_linkage_gate 项。
    """
    obj_by_id = {o["id"]: o for o in s2_objs}
    obj_ids_referenced: set[str] = set()

    missing_obj_id: list[str] = []
    invalid_obj_id: list[str] = []
    desc_mismatch: list[dict[str, str]] = []

    for tc in test_cases:
        tc_id = tc.get("case_id", "<unknown>")
        obj_id = tc.get("obj_id")
        if not obj_id:
            missing_obj_id.append(tc_id)
            continue
        if obj_id not in obj_by_id:
            invalid_obj_id.append(tc_id)
            continue
        obj_ids_referenced.add(obj_id)
        expected_name = obj_by_id[obj_id].get("obj_name", "")
        actual_desc = tc.get("用例描述", "")
        if actual_desc != expected_name:
            desc_mismatch.append({
                "tc": tc_id,
                "obj_id": obj_id,
                "用例描述": actual_desc,
                "expected_obj_name": expected_name,
            })

    total_objs = len(obj_by_id)
    coverage = (len(obj_ids_referenced) / total_objs) if total_objs else 1.0
    unreferenced = sorted(set(obj_by_id) - obj_ids_referenced)

    return {
        "passed": (
            not missing_obj_id
            and not invalid_obj_id
            and not desc_mismatch
            and coverage >= 1.0
        ),
        "missing_obj_id": missing_obj_id,
        "invalid_obj_id": invalid_obj_id,
        "desc_mismatch": desc_mismatch,
        "obj_linkage_coverage": round(coverage, 4),
        "total_s2_objs": total_objs,
        "referenced_s2_objs": sorted(obj_ids_referenced),
        "unreferenced_s2_objs": unreferenced,
    }


# ─────────────────────────────────────────────────────────────────────────────
# v11: S5 TP → S6 TC → S2 OBJ+FP 双层覆盖门禁
# ─────────────────────────────────────────────────────────────────────────────


def _load_s5_test_points(req_name: str, version: str) -> list[dict] | None:
    """加载 S5 测试点列表——v11 TC→TP 链路校验的输入源。

    返回 None 表示无法加载（目录不存在 / 文件缺失 / JSON 解析失败）。
    """
    assets_root = _ROOT / "workflow_assets" / req_name / version / "「S5 测试点生成」"
    tp_path = assets_root / "test_points.json"
    if not tp_path.exists():
        return None
    try:
        data = json.loads(tp_path.read_text(encoding="utf-8"))
        # S5 文件可能存为 scenario_test_points 或 test_points（两种格式兼容）
        tps = data.get("scenario_test_points") or data.get("test_points") or []
        return tps if isinstance(tps, list) else None
    except Exception:
        return None


def _validate_obj_fp_linkage(
    test_cases: list[dict],
    s5_test_points: list[dict] | None,
    s2_objs: list[dict],
) -> dict[str, Any]:
    """S6 TC + S5 TP + S2 OBJ+FP 双层覆盖校验（v11 §FP 全覆盖 L3 兜底）。

    门禁（违反任一 → passed=False）：
    - 每个 TC 有 s5_ref（缺失 → missing_s5_ref 列出 case_id）
    - 每个 s5_ref 在 S5 TP 列表中存在（缺失 → invalid_s5_ref）
    - obj_linkage_coverage ≥ 1.0（沿用 v10）
    - fp_linkage_coverage ≥ 1.0（以 s5_ref 推导的 FP 集合为准）
    - TC.obj_id == s5_ref 推导出的 OBJ.id（冗余校验）

    返回结构供 postflight_gate.json 写入 obj_fp_linkage_gate 项。
    """
    obj_by_id = {o["id"]: o for o in s2_objs}
    all_fp_ids: set[str] = set()
    fp_to_obj: dict[str, str] = {}
    for o in s2_objs:
        for fp in o.get("feature_points", []):
            fp_id = fp["id"]
            all_fp_ids.add(fp_id)
            fp_to_obj[fp_id] = o["id"]

    tp_by_id: dict[str, dict] = {}
    if s5_test_points:
        tp_by_id = {tp.get("tp_id"): tp for tp in s5_test_points if tp.get("tp_id")}

    obj_ids_referenced: set[str] = set()
    fp_ids_referenced: set[str] = set()
    missing_s5_ref: list[str] = []
    invalid_s5_ref: list[str] = []
    obj_mismatch_via_s5_ref: list[dict[str, str]] = []

    for tc in test_cases:
        tc_id = tc.get("case_id", "<unknown>")
        s5_ref = tc.get("s5_ref")
        obj_id = tc.get("obj_id")

        # OBJ 维度（沿用 v10 风格——即使 s5_ref 缺也校验 obj_id）
        if obj_id and obj_id in obj_by_id:
            obj_ids_referenced.add(obj_id)

        # s5_ref 必填
        if not s5_ref:
            missing_s5_ref.append(tc_id)
            continue
        if s5_ref not in tp_by_id:
            invalid_s5_ref.append(tc_id)
            continue

        # 从 s5_ref 推导 FP
        tp = tp_by_id[s5_ref]
        fp_ref = tp.get("feature_point_ref") or tp.get("feature_point_id")
        if not fp_ref:
            invalid_s5_ref.append(tc_id)
            continue
        if fp_ref not in all_fp_ids:
            invalid_s5_ref.append(tc_id)
            continue

        fp_ids_referenced.add(fp_ref)

        # TC.obj_id 必须 == s5_ref 推导出的 OBJ.id（冗余校验）
        derived_obj = fp_to_obj.get(fp_ref)
        if derived_obj and obj_id and obj_id != derived_obj:
            obj_mismatch_via_s5_ref.append({
                "tc": tc_id,
                "tc_obj_id": obj_id,
                "s5_ref": s5_ref,
                "derived_obj_id": derived_obj,
            })

    total_objs = len(obj_by_id)
    total_fps = len(all_fp_ids)
    obj_coverage = (len(obj_ids_referenced) / total_objs) if total_objs else 1.0
    fp_coverage = (len(fp_ids_referenced) / total_fps) if total_fps else 1.0
    unreferenced_objs = sorted(set(obj_by_id) - obj_ids_referenced)
    unreferenced_fps = sorted(all_fp_ids - fp_ids_referenced)

    passed = (
        not missing_s5_ref
        and not invalid_s5_ref
        and not obj_mismatch_via_s5_ref
        and obj_coverage >= 1.0
        and fp_coverage >= 1.0
    )

    return {
        "passed": passed,
        "obj_linkage_coverage": round(obj_coverage, 4),
        "fp_linkage_coverage": round(fp_coverage, 4),
        "total_s2_objs": total_objs,
        "total_s2_fps": total_fps,
        "referenced_s2_objs": sorted(obj_ids_referenced),
        "referenced_s2_fps": sorted(fp_ids_referenced),
        "unreferenced_s2_objs": unreferenced_objs,
        "unreferenced_s2_fps": unreferenced_fps,
        "missing_s5_ref": missing_s5_ref,
        "invalid_s5_ref": invalid_s5_ref,
        "obj_mismatch_via_s5_ref": obj_mismatch_via_s5_ref,
    }


def run_obj_fp_linkage_validation(
    req_name: str,
    version: str,
    output_dir: Path | str,
) -> dict[str, Any]:
    """一站式：加载 TC + S5 TP + S2 OBJ + 跑 _validate_obj_fp_linkage + 写 postflight_gate。

    Phase 5.2 接入点。Phase 5.4 由 S6 调用。
    """
    if isinstance(output_dir, str):
        output_dir = Path(output_dir)

    tc_path = output_dir / "test_cases.json"
    if not tc_path.exists():
        return {"passed": False, "error": f"test_cases.json not found at {tc_path}"}

    s2_objs = _load_s2_requirement_objects(req_name, version)
    if s2_objs is None:
        return {"passed": False, "error": f"S2 requirement_objects.json not loadable for {req_name}/{version}"}

    s5_tps = _load_s5_test_points(req_name, version)
    if s5_tps is None:
        return {"passed": False, "error": f"S5 test_points.json not loadable for {req_name}/{version}", "hint": "需要 S5 提供 TP 列表"}

    try:
        tc_data = json.loads(tc_path.read_text(encoding="utf-8"))
        cases = tc_data.get("test_cases", [])
    except Exception as e:
        return {"passed": False, "error": f"Failed to parse test_cases.json: {e}"}

    report = _validate_obj_fp_linkage(cases, s5_tps, s2_objs)

    gate_path = output_dir / "postflight_gate.json"
    gate: dict[str, Any] = {}
    if gate_path.exists():
        try:
            gate = json.loads(gate_path.read_text(encoding="utf-8"))
        except Exception:
            gate = {}
    if "checks" not in gate:
        gate["checks"] = {}
    gate["checks"]["obj_fp_linkage_gate"] = report
    gate["checked_at"] = datetime.now().isoformat(timespec="seconds")
    # obj_fp_linkage 是硬门禁——直接写入顶层 passed
    if not report["passed"]:
        gate["passed"] = False
    gate_path.write_text(
        json.dumps(gate, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return report


def _self_test_obj_fp_linkage() -> int:
    """v11 §9.1.1 self-test 豁免条款应用：本函数是 v11 双层校验的 self-test。

    验证 4 场景：
    1. 完整通过：所有 TC 有 s5_ref → 推导 FP 全覆盖 OBJ 全覆盖
    2. missing_s5_ref：1 个 TC 缺 s5_ref
    3. invalid_s5_ref：1 个 TC 的 s5_ref 不在 TP 列表
    4. fp_coverage < 1.0：S2 有 3 个 FP 但 TC 只覆盖 2 个
    """
    s2_objs = [
        {
            "id": "TEST-OBJ-01",
            "obj_name": "对象A",
            "feature_points": [
                {"id": "TEST-OBJ-01-FP-1", "fp_desc": "FP-1 desc"},
                {"id": "TEST-OBJ-01-FP-2", "fp_desc": "FP-2 desc"},
            ],
        },
        {
            "id": "TEST-OBJ-02",
            "obj_name": "对象B",
            "feature_points": [
                {"id": "TEST-OBJ-02-FP-1", "fp_desc": "FP-1 desc"},
            ],
        },
    ]
    s5_tps = [
        {"tp_id": "TP-1", "feature_point_ref": "TEST-OBJ-01-FP-1"},
        {"tp_id": "TP-2", "feature_point_ref": "TEST-OBJ-01-FP-2"},
        {"tp_id": "TP-3", "feature_point_ref": "TEST-OBJ-02-FP-1"},
    ]

    # 场景 1: 完整通过
    r1 = _validate_obj_fp_linkage(
        [
            {"case_id": "TC-1", "s5_ref": "TP-1", "obj_id": "TEST-OBJ-01"},
            {"case_id": "TC-2", "s5_ref": "TP-2", "obj_id": "TEST-OBJ-01"},
            {"case_id": "TC-3", "s5_ref": "TP-3", "obj_id": "TEST-OBJ-02"},
        ],
        s5_tps,
        s2_objs,
    )
    assert r1["passed"], f"场景 1 应通过，实际: {r1}"
    assert r1["obj_linkage_coverage"] == 1.0
    assert r1["fp_linkage_coverage"] == 1.0

    # 场景 2: missing_s5_ref
    r2 = _validate_obj_fp_linkage(
        [
            {"case_id": "TC-1", "s5_ref": "TP-1", "obj_id": "TEST-OBJ-01"},
            {"case_id": "TC-2", "obj_id": "TEST-OBJ-01"},  # 缺 s5_ref
        ],
        s5_tps,
        s2_objs,
    )
    assert not r2["passed"], "场景 2 应失败"
    assert "TC-2" in r2["missing_s5_ref"]

    # 场景 3: invalid_s5_ref
    r3 = _validate_obj_fp_linkage(
        [
            {"case_id": "TC-1", "s5_ref": "TP-1", "obj_id": "TEST-OBJ-01"},
            {"case_id": "TC-2", "s5_ref": "TP-DOES-NOT-EXIST", "obj_id": "TEST-OBJ-01"},
        ],
        s5_tps,
        s2_objs,
    )
    assert not r3["passed"], "场景 3 应失败"
    assert "TC-2" in r3["invalid_s5_ref"]

    # 场景 4: fp_coverage < 1.0（只覆盖 2 个 FP 中的 1 个）
    r4 = _validate_obj_fp_linkage(
        [
            {"case_id": "TC-1", "s5_ref": "TP-1", "obj_id": "TEST-OBJ-01"},
        ],
        s5_tps,
        s2_objs,
    )
    assert not r4["passed"], "场景 4 应失败"
    assert abs(r4["fp_linkage_coverage"] - 1/3) < 0.001, f"场景 4 fp_coverage 应 ≈ 1/3，实际={r4['fp_linkage_coverage']}"
    assert len(r4["unreferenced_s2_fps"]) == 2

    print("✅ _validate_obj_fp_linkage self-test: 4 场景全过")
    return 0


def run_obj_linkage_validation(
    req_name: str,
    version: str,
    output_dir: Path | str,
) -> dict[str, Any]:
    """一站式：加载 TC + S5 TP + S2 OBJ + 跑 _validate_obj_fp_linkage + 写 postflight_gate。

    Phase 5.2 接入点。Phase 5.4 由 S6 调用。
    """
    if isinstance(output_dir, str):
        output_dir = Path(output_dir)

    tc_path = output_dir / "test_cases.json"
    if not tc_path.exists():
        return {"passed": False, "error": f"test_cases.json not found at {tc_path}"}

    s2_objs = _load_s2_requirement_objects(req_name, version)
    if s2_objs is None:
        return {"passed": False, "error": f"S2 requirement_objects.json not loadable for {req_name}/{version}"}

    try:
        tc_data = json.loads(tc_path.read_text(encoding="utf-8"))
        cases = tc_data.get("test_cases", [])
    except Exception as e:
        return {"passed": False, "error": f"Failed to parse test_cases.json: {e}"}

    report = _validate_obj_linkage(cases, s2_objs)

    gate_path = output_dir / "postflight_gate.json"
    gate: dict[str, Any] = {}
    if gate_path.exists():
        try:
            gate = json.loads(gate_path.read_text(encoding="utf-8"))
        except Exception:
            gate = {}
    if "checks" not in gate:
        gate["checks"] = {}
    gate["checks"]["obj_linkage_gate"] = report
    gate["passed"] = bool(gate.get("passed", False)) and bool(report["passed"])
    gate["checked_at"] = datetime.now().isoformat(timespec="seconds")
    gate_path.write_text(
        json.dumps(gate, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return report


def _self_test_obj_linkage() -> int:
    """v10 §9.1.1 self-test 豁免条款应用：本函数是新增的 self-test，不改任何业务函数签名。

    验证 4 个场景：
    1. 完整通过：所有 TC 有 obj_id + 用例描述 == obj_name + 覆盖率 100%
    2. missing_obj_id：1 个 TC 缺 obj_id
    3. desc_mismatch：1 个 TC 的用例描述 != obj_name
    4. coverage < 1.0：S2 有 2 个 OBJ 但只引用 1 个
    """
    s2_objs = [
        {"id": "TEST-OBJ-01", "obj_name": "对象A"},
        {"id": "TEST-OBJ-02", "obj_name": "对象B"},
    ]
    # 场景 1: 完整通过
    r1 = _validate_obj_linkage(
        [
            {"case_id": "TC-1", "obj_id": "TEST-OBJ-01", "用例描述": "对象A"},
            {"case_id": "TC-2", "obj_id": "TEST-OBJ-02", "用例描述": "对象B"},
        ],
        s2_objs,
    )
    assert r1["passed"], f"场景 1 应通过，实际: {r1}"
    assert r1["obj_linkage_coverage"] == 1.0

    # 场景 2: missing_obj_id
    r2 = _validate_obj_linkage(
        [
            {"case_id": "TC-1", "obj_id": "TEST-OBJ-01", "用例描述": "对象A"},
            {"case_id": "TC-2", "用例描述": "对象B"},  # 缺 obj_id
        ],
        s2_objs,
    )
    assert not r2["passed"], "场景 2 应失败"
    assert "TC-2" in r2["missing_obj_id"]

    # 场景 3: desc_mismatch
    r3 = _validate_obj_linkage(
        [
            {"case_id": "TC-1", "obj_id": "TEST-OBJ-01", "用例描述": "对象A"},
            {"case_id": "TC-2", "obj_id": "TEST-OBJ-02", "用例描述": "对象X"},  # mismatch
        ],
        s2_objs,
    )
    assert not r3["passed"], "场景 3 应失败"
    assert any(m["tc"] == "TC-2" for m in r3["desc_mismatch"])

    # 场景 4: coverage < 1.0（只引用 1 个 OBJ，另一个未引用）
    r4 = _validate_obj_linkage(
        [
            {"case_id": "TC-1", "obj_id": "TEST-OBJ-01", "用例描述": "对象A"},
        ],
        s2_objs,
    )
    assert not r4["passed"], "场景 4 应失败"
    assert r4["obj_linkage_coverage"] == 0.5
    assert "TEST-OBJ-02" in r4["unreferenced_s2_objs"]

    print("✅ _validate_obj_linkage self-test: 4 场景全过")
    return 0


def _self_test_xlsx_cli() -> int:
    """Verify three JSON schemas, immutable public headers, and two-sheet routing."""
    import tempfile
    import openpyxl

    samples = [
        [
            {"case_id": "UI-TC-001", "module": "UI", "用例状态": "Ready"},
            {"case_id": "UI-TC-002", "module": "UI", "用例状态": "Draft"},
            {"case_id": "UI-TC-003", "module": "UI", "用例状态": "Rejected"},
        ],
        {"test_cases": [
            {"case_id": "BIZ-TC-001", "module": "BIZ", "用例状态": "Ready"},
            {"case_id": "BIZ-TC-002", "module": "BIZ", "用例状态": "Draft"},
        ]},
        {"meta": {"stage": "S6"}, "test_cases": [
            {"case_id": "LOG-TC-001", "module": "LOG", "用例状态": "Ready"},
            {"case_id": "LOG-TC-002", "module": "LOG", "用例状态": "Rejected"},
        ]},
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        root = Path(tmpdir)
        for index, sample in enumerate(samples, start=1):
            json_path = root / f"schema_{index}.json"
            xlsx_path = root / f"schema_{index}.xlsx"
            json_path.write_text(json.dumps(sample, ensure_ascii=False), encoding="utf-8")
            export_test_cases_json_to_xlsx(json_path, xlsx_path)
            workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
            assert workbook.sheetnames == ["测试用例", "Draft-Rejected附录"]
            for worksheet in workbook.worksheets:
                headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
                assert headers == _XLSX_HEADERS_V3
            main_statuses = [row[8] for row in workbook["测试用例"].iter_rows(min_row=2, values_only=True)]
            appendix_statuses = [
                row[8]
                for row in workbook["Draft-Rejected附录"].iter_rows(min_row=2, values_only=True)
            ]
            assert main_statuses and set(main_statuses) == {"Ready"}
            assert set(appendix_statuses).issubset({"Draft", "Rejected"})
            workbook.close()

    print("test_case_formatter xlsx CLI self-test: PASS")
    return 0


def _self_test_round17_obj_grouping() -> int:
    """Cases 7-9: verify OBJ merges, no spacer rows, sorting, and color bands."""
    import tempfile
    import openpyxl

    cases = [
        {
            "case_id": "BIZ-TC-002",
            "module": "BIZ",
            "用例描述": "对象B",
            "obj_id": "OBJ-B",
            "feature_point_ref": "OBJ-B-FP-2",
            "用例状态": "Ready",
            "功能描述": "功能B",
            "前置条件": "前置B",
        },
        {
            "case_id": "BIZ-TC-001",
            "module": "BIZ",
            "用例描述": "对象B",
            "obj_id": "OBJ-B",
            "feature_point_ref": "OBJ-B-FP-1",
            "用例状态": "Ready",
            "功能描述": "功能B",
            "前置条件": "前置B",
        },
        {
            "case_id": "UI-TC-003",
            "module": "UI",
            "用例描述": "对象A",
            "obj_id": "OBJ-A",
            "feature_point_ref": "OBJ-A-FP-1",
            "用例状态": "Ready",
            "功能描述": "功能A",
            "前置条件": "前置A",
        },
        {
            "case_id": "UI-TC-001",
            "module": "UI",
            "用例描述": "对象A",
            "obj_id": "OBJ-A",
            "feature_point_ref": "OBJ-A-FP-1",
            "用例状态": "Ready",
            "功能描述": "功能A",
            "前置条件": "前置A",
        },
    ]
    sort_options = {
        "by": ["obj_id", "case_id"],
        "obj_color_band": True,
        # Round 18: 5-col request (per user report R18-B1); _is_column_uniform
        # filter protects against data loss when E/F are not uniform.
        "merge_obj_headers": ["用例描述", "所属模块", "关联需求", "功能描述", "前置条件"],
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "round17.xlsx"
        _save_xlsx(
            cases,
            Path(tmpdir),
            export_profile=copy.deepcopy(_ROUND17_XLSX_PROFILE),
            output_path=xlsx_path,
            sort_options=sort_options,
            description_index_sheet=True,
        )
        workbook = openpyxl.load_workbook(xlsx_path)
        worksheet = workbook["测试用例"]

        # Case 7: two OBJ blocks × five shared columns (all uniform in test data).
        # Round 18 expectation: 10 merge regions (2 OBJ × 5 cols).
        assert len(worksheet.merged_cells.ranges) == 10, worksheet.merged_cells.ranges
        assert {str(cell_range) for cell_range in worksheet.merged_cells.ranges} == {
            "B2:B3", "C2:C3", "D2:D3", "E2:E3", "F2:F3",
            "B4:B5", "C4:C5", "D4:D5", "E4:E5", "F4:F5",
        }

        # Case 8: no empty rows below the header.
        blank_rows = [
            row_idx
            for row_idx in range(2, worksheet.max_row + 1)
            if all(
                worksheet.cell(row=row_idx, column=col_idx).value in (None, "")
                for col_idx in range(1, worksheet.max_column + 1)
            )
        ]
        assert worksheet.max_row == len(cases) + 1, worksheet.max_row
        assert blank_rows == [], blank_rows

        # Case 9: obj_id ascending, then case_id ascending; OBJ colors differ.
        assert [worksheet.cell(row=row_idx, column=1).value for row_idx in range(2, 6)] == [
            "UI-TC-001", "UI-TC-003", "BIZ-TC-001", "BIZ-TC-002"
        ]
        assert worksheet["A2"].fill.fgColor.rgb == worksheet["A3"].fill.fgColor.rgb
        assert worksheet["A4"].fill.fgColor.rgb == worksheet["A5"].fill.fgColor.rgb
        assert worksheet["A2"].fill.fgColor.rgb != worksheet["A4"].fill.fgColor.rgb
        workbook.close()

    # Case 10: 5-col merge attempted, but E/F non-uniform → guarded no-merge.
    # Synthesizes an OBJ where 用例描述/所属模块/关联需求 are uniform but
    # 功能描述/前置条件 vary per TC. The merge filter must skip E/F to avoid
    # data loss, leaving only 3 merges per OBJ.
    mixed_cases = [
        {
            "case_id": "TC-M01",
            "module": "BIZ",
            "用例描述": "混合OBJ",
            "obj_id": "OBJ-MIX",
            "用例状态": "Ready",
            "功能描述": "功能-A",
            "前置条件": "前置-X",
        },
        {
            "case_id": "TC-M02",
            "module": "BIZ",
            "用例描述": "混合OBJ",
            "obj_id": "OBJ-MIX",
            "用例状态": "Ready",
            "功能描述": "功能-B",   # varies → E must not merge
            "前置条件": "前置-X",
        },
        {
            "case_id": "TC-M03",
            "module": "BIZ",
            "用例描述": "混合OBJ",
            "obj_id": "OBJ-MIX",
            "用例状态": "Ready",
            "功能描述": "功能-A",
            "前置条件": "前置-Y",   # varies → F must not merge
        },
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "round18_mixed.xlsx"
        _save_xlsx(
            mixed_cases,
            Path(tmpdir),
            export_profile=copy.deepcopy(_ROUND17_XLSX_PROFILE),
            output_path=xlsx_path,
            sort_options={
                "by": ["obj_id", "case_id"],
                "obj_color_band": True,
                "merge_obj_headers": ["用例描述", "所属模块", "关联需求", "功能描述", "前置条件"],
            },
            description_index_sheet=False,
        )
        workbook = openpyxl.load_workbook(xlsx_path)
        worksheet = workbook["测试用例"]
        merge_ranges = {str(r) for r in worksheet.merged_cells.ranges}
        # Only B/C/D merge (uniform), E and F are protected by _is_column_uniform.
        assert merge_ranges == {"B2:B4", "C2:C4", "D2:D4"}, merge_ranges
        # Verify E and F are NOT merged but each row carries its own value.
        assert worksheet["E2"].value == "功能-A"
        assert worksheet["E3"].value == "功能-B"
        assert worksheet["E4"].value == "功能-A"
        assert worksheet["F2"].value == "前置-X"
        assert worksheet["F3"].value == "前置-X"
        assert worksheet["F4"].value == "前置-Y"
        workbook.close()

    # Case 11: H column (预期结果) is never merged, regardless of OBJ block size.
    # Each H cell carries its own string (which may contain \n-separated lines
    # from merge_grouped_inplace concatenation, but those are 1 TC's content,
    # not a visual merge artifact).
    h_cases = [
        {
            "case_id": "TC-H01",
            "module": "BIZ",
            "用例描述": "OBJ-X",
            "obj_id": "OBJ-X",
            "用例状态": "Ready",
            "expected_results": ["预期行1", "预期行2"],
            "steps": [{"step_num": 1, "action": "动作1"}],
            "preconditions": ["前置"],
        },
        {
            "case_id": "TC-H02",
            "module": "BIZ",
            "用例描述": "OBJ-X",
            "obj_id": "OBJ-X",
            "用例状态": "Ready",
            "expected_results": ["预期单行"],
            "steps": [{"step_num": 1, "action": "动作2"}],
            "preconditions": ["前置"],
        },
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "round18_h.xlsx"
        _save_xlsx(
            h_cases,
            Path(tmpdir),
            export_profile=copy.deepcopy(_ROUND17_XLSX_PROFILE),
            output_path=xlsx_path,
            sort_options={
                "by": ["obj_id", "case_id"],
                "obj_color_band": True,
                "merge_obj_headers": ["用例描述", "所属模块", "关联需求", "功能描述", "前置条件"],
            },
            description_index_sheet=False,
        )
        workbook = openpyxl.load_workbook(xlsx_path)
        worksheet = workbook["测试用例"]
        merge_cols = {r.min_col for r in worksheet.merged_cells.ranges}
        # Header order: A=用例ID, B=所属模块, C=用例描述, D=关联需求,
        # E=功能描述, F=前置条件, G=操作步骤, H=预期结果, I=优先级, J=用例状态,
        # K=备注. Column H = 8 must NEVER appear in any merge range.
        assert 8 not in merge_cols, (
            f"H column unexpectedly merged: {worksheet.merged_cells.ranges}"
        )
        # Verify each H cell is independent (not concat'd across rows).
        h2 = worksheet["H2"].value
        h3 = worksheet["H3"].value
        assert h2 is not None and "\n" in h2, h2
        assert h3 == "预期单行", h3
        workbook.close()

    print("test_case_formatter Round 17/18 cases 7-11: PASS")
    return 0


def self_test() -> int:
    """Run formatter linkage tests and public XLSX tests."""
    v10_rc = _self_test_obj_linkage()
    v11_rc = _self_test_obj_fp_linkage()
    xlsx_rc = _self_test_xlsx_cli()
    round17_rc = _self_test_round17_obj_grouping()
    return 0 if all(rc == 0 for rc in (v10_rc, v11_rc, xlsx_rc, round17_rc)) else 1


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--self-test":
        sys.exit(self_test())
    sys.exit(_cli_migrate_modules(sys.argv[1:]))
