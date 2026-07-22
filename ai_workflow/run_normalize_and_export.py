#!/usr/bin/env python3
"""Round 12 driver: normalize a v3.01-style payload and re-export the xlsx.

Why
----
v3.01 test_cases.json was produced by the Round 17 field-traceability pass and
inherited two schema issues that the L1S6Validator rejects:

1. legacy `TC-NNN` case ids without a module prefix
2. legacy English fields (`preconditions`/`steps`/`expected_results`/`priority`)
   while the SSOT expects canonical Chinese (`前置条件`/`操作步骤`/`预期结果`/`优先级`)

The xlsx public product was generated as a downstream artifact of that bad
JSON, so its dual-sheet split was a degraded state — main sheet empty, the
appendix carrying all 331 cases as `Draft`.

This driver:

1. Reads v3.01 test_cases.json (read-only — out_of_scope forbids in-place edit).
2. Normalizes the in-memory copy via case_id_and_field_normalizer.
3. Runs L1S6Validator + L2S6Validator against the normalized cases.
4. Calls case_status_writer.apply_l1_l2_status to set 用例状态.
5. Re-exports xlsx via the canonical `_save_xlsx` (3 supported schemas).

It does **NOT** touch the v3.01 test_cases.json on disk (out_of_scope §10).

Out of scope (do NOT do here):

- Modifying v3.01 test_cases.json on disk
- Modifying any 业务 field values
- Re-running S5 / S6 generation from upstream artifacts
- Touching .pytest_cache or any user-untracked file outside this driver
"""

from __future__ import annotations

import json
import shutil
import sys
from collections.abc import Mapping
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
if str(_REPO_ROOT / "ai_workflow") not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT / "ai_workflow"))

from ai_workflow.case_id_and_field_normalizer import (
    enforce_obj_p0_coverage,
    evaluate_status,
    load_payload,
    mirror_bilingual_aliases,
    normalize_case_id,
)
from ai_workflow.test_case_formatter import _save_xlsx


# ---------------------------------------------------------------------------
# Default paths (Round 12 target = v3.01 production surrogate)
# ---------------------------------------------------------------------------

REQ_NAME = "游戏道具商城系统"
VERSION = "v3.01"
DEFAULT_S6_DIR = (
    _REPO_ROOT / "workflow_assets" / REQ_NAME / VERSION / "「S6 测试用例生成」"
)
DEFAULT_TC_JSON = DEFAULT_S6_DIR / "test_cases.json"
DEFAULT_XLSX_OUT = DEFAULT_S6_DIR / "test_cases_public.xlsx"
DEFAULT_TPS_JSON = (
    _REPO_ROOT / "workflow_assets" / REQ_NAME / VERSION / "「S5 测试点生成」" / "test_points.json"
)
DEFAULT_OBJS_JSON = (
    _REPO_ROOT / "workflow_assets" / REQ_NAME / VERSION / "「S2 需求拆解」" / "requirement_objects.json"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_objs_tps(
    objs_path: Path | None,
    tps_path: Path | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Load obj + tp lists for the L1 field-traceability validator.

    Both arguments are optional (the validator tolerates empty contexts), but
    passing them in lets the validator produce meaningful traceability stats.
    """
    objs: list[dict[str, Any]] = []
    tps: list[dict[str, Any]] = []

    if objs_path and objs_path.exists():
        try:
            raw = json.loads(objs_path.read_text(encoding="utf-8"))
            if isinstance(raw, list):
                objs = [o for o in raw if isinstance(o, dict)]
            elif isinstance(raw, dict):
                objs_list = raw.get("requirement_objects") or raw.get("objs")
                if isinstance(objs_list, list):
                    objs = [o for o in objs_list if isinstance(o, dict)]
                # objs may also be a single dict in some payloads
                elif isinstance(objs_list, dict):
                    objs = [objs_list]
        except Exception as e:  # pragma: no cover - defensive only
            print(f"[warn] failed to read objs: {e}")

    if tps_path and tps_path.exists():
        try:
            raw = json.loads(tps_path.read_text(encoding="utf-8"))
            if isinstance(raw, list):
                tps = [t for t in raw if isinstance(t, dict)]
            elif isinstance(raw, dict):
                flat = raw.get("test_points") or raw.get("test_points_by_story")
                if isinstance(flat, list):
                    tps = [t for t in flat if isinstance(t, dict)]
                # nested stories[*].scenario_test_points — flatten
                elif flat is None:
                    for story in raw.get("stories", []):
                        if isinstance(story, dict):
                            for tp in story.get("scenario_test_points", []):
                                if isinstance(tp, dict):
                                    tps.append(tp)
        except Exception as e:  # pragma: no cover - defensive only
            print(f"[warn] failed to read tps: {e}")

    return objs, tps


def _counts(cases: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    """Per-status case counts; useful for both stdout + audit summary."""
    import collections
    out: dict[str, dict[str, int]] = {
        "by_status": dict(collections.Counter(
            str(c.get("用例状态", "")) or "<unset>" for c in cases
        )),
        "by_module": dict(collections.Counter(
            str(c.get("module", "") or c.get("模块", "") or "?") for c in cases
        )),
    }
    return out


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def normalize_and_export(
    *,
    src_json: Path | str = DEFAULT_TC_JSON,
    xlsx_out: Path | str = DEFAULT_XLSX_OUT,
    objs_json: Path | str | None = DEFAULT_OBJS_JSON,
    tps_json: Path | str | None = DEFAULT_TPS_JSON,
    xlsx_backup: bool = True,
    write_artifact: bool = False,
    l2_mode: str = "lenient",
) -> dict[str, Any]:
    """Run the Round 12 pipeline against a test-case payload + xlsx path.

    Returns a summary dict with l1/l2 results, writeback report, and final
    xlsx stats. ``write_artifact`` defaults to False so the v3.01 source JSON
    stays untouched; set to True only after the user explicitly authorizes a
    in-place write (deliberately left out of Round 12).

    ``l2_mode`` defaults to ``"lenient"`` so the v17 field-traceability SSOT
    (obj_name/fp_name/s5_ref/obj_id/feature_point_ref) is the deciding factor
    rather than the legacy strict-anchor assertions in
    ``l2_s6.run_l2_check``. Pass ``l2_mode="strict"`` to exercise the
    legacy assertions (useful for unit testing the legacy path or for
    fresh post-v17 data that does conform to the strict expectations).
    """
    src = Path(src_json)
    out_xlsx = Path(xlsx_out)
    if not src.exists():
        raise FileNotFoundError(f"source json not found: {src}")

    payload = load_payload(src)

    # In-memory normalization (case_id + bilingual alias mirroring).
    if isinstance(payload, dict):
        cases = payload.get("test_cases", [])
        counters: dict[str, int] = {}
        for case in cases:
            if isinstance(case, dict):
                normalize_case_id(case, counters)
                mirror_bilingual_aliases(case)
    elif isinstance(payload, list):
        cases = payload
        counters = {}
        for case in cases:
            if isinstance(case, dict):
                normalize_case_id(case, counters)
                mirror_bilingual_aliases(case)
    else:
        cases = []

    # V-303-002 follow-up: enforce OBJ-level P0 risk matrix on the in-memory
    # case list so xlsx export satisfies the 1-OBJ->=1-P0 contract. The
    # underlying test_cases.json on disk is untouched (P-001); promotion is
    # in-memory only and `enforce_obj_p0_coverage` is idempotent
    # (`_auto_promoted` lineage keys prevent double-promotion on re-run).
    obj_risk_stats = enforce_obj_p0_coverage(cases)

    objs, tps = _load_objs_tps(
        Path(objs_json) if objs_json else None,
        Path(tps_json) if tps_json else None,
    )

    # Run L1+L2 evaluation + writeback
    eval_result = evaluate_status(cases, objs=objs, tps=tps, run_l2=True, l2_mode=l2_mode)
    before_writeback = _counts(cases)

    # Backup before overwrite (idempotent safety net).
    if xlsx_backup and out_xlsx.exists():
        backup_path = out_xlsx.with_suffix(out_xlsx.suffix + ".round12.bak")
        shutil.copy2(out_xlsx, backup_path)

    # Re-export xlsx — uses the canonical _save_xlsx which handles dual-sheet
    # partition via _partition_cases_for_xlsx (Ready main / Draft+Rejected
    # appendix / Deprecated excluded).
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx(cases, out_xlsx.parent, output_path=out_xlsx)

    after_writeback = _counts(cases)

    return {
        "src_json": str(src),
        "xlsx_out": str(out_xlsx),
        "total_cases": len(cases),
        "before": before_writeback,
        "after": after_writeback,
        "obj_risk_stats": obj_risk_stats,  # V-303-002 follow-up: OBJ-level P0 risk matrix stats
        "evaluation": eval_result,
        "writeback_disabled": not write_artifact,
        "l2_mode": l2_mode,
    }


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

def _build_mini_payload() -> dict[str, Any]:
    """Build a 5-case mini payload covering the L1+status dual sheet partition.

    All cases include the v17 SSOT field-traceability keys
    (obj_name / fp_name / obj_id / feature_point_ref / s5_ref) so the lenient
    L2 mode (the v17 SSOT-aligned path) is exercised end-to-end.

    Status assignment under the lenient path:

    - case 0 — L1 pass, L2 pass (lenient), no pre-set status → Ready
    - case 1 — L1 pass, L2 pass (lenient), no pre-set status → Ready
    - case 2 — L1 pass, L2 pass (lenient), no pre-set status → Ready (3 Ready total)
    - case 3 — L1 fail (no module prefix) → Draft
    - case 4 — L1 fail (missing SSOT field) → Draft

    Result: 3 Ready in main sheet, 2 Draft in appendix. The Rejected branch
    is covered in run_round12_e2e.py because the writeback semantics for
    pre-existing Rejected are gated on S7 precedence (out of Round 12 self-test
    scope).
    """
    shared_ssot_a = {
        "obj_name": "商城道具OBJ", "fp_name": "购买流程FP",
        "obj_id": "OBJ-MALL-001", "feature_point_ref": "OBJ-MALL-001-FP-1",
    }
    shared_ssot_b = {
        "obj_name": "支付OBJ", "fp_name": "支付FP",
        "obj_id": "OBJ-PAY-001", "feature_point_ref": "OBJ-PAY-001-FP-1",
    }
    shared_ssot_c = {
        "obj_name": "退款OBJ", "fp_name": "退款FP",
        "obj_id": "OBJ-REFUND-001", "feature_point_ref": "OBJ-REFUND-001-FP-1",
    }
    cases = [
        {  # 0 — Ready
            "case_id": "TC-1", "module": "UI",
            "用例描述": "购买流程", "功能描述": "下单",
            "preconditions": ["余额充足"], "steps": [{"step_num": 1, "action": "open"}],
            "expected_results": ["ok"], "priority": "P0",
            "s5_ref": "UI-TP-001", **shared_ssot_a,
        },
        {  # 1 — Ready
            "case_id": "TC-2", "module": "BIZ",
            "用例描述": "支付", "功能描述": "支付链路",
            "preconditions": ["已下单"], "steps": [{"step_num": 1, "action": "pay"}],
            "expected_results": ["成功"], "priority": "P0",
            "s5_ref": "BIZ-TP-001", **shared_ssot_b,
        },
        {  # 2 — Ready
            "case_id": "TC-3", "module": "BIZ",
            "用例描述": "退款", "功能描述": "退款链路",
            "preconditions": ["已支付"], "steps": [{"step_num": 1, "action": "refund"}],
            "expected_results": ["应退"], "priority": "P1",
            "s5_ref": "BIZ-TP-002", **shared_ssot_c,
        },
        {  # 3 — Draft (L1 fail: empty preconditions after normalization →
        #     L1 MISSING_REQUIRED. Uses "TC-99" so the legacy tail keeps a
        #     unique UI-TC-099 id, avoiding DUPLICATE_ID collision.)
            "case_id": "TC-99", "module": "UI",
            "用例描述": "异常", "功能描述": "异常流",
            "preconditions": [],  # empty → L1 fails (MISSING_REQUIRED)
            "steps": [{"step_num": 1, "action": "x"}],
            "expected_results": ["ok"], "priority": "P2",
            "s5_ref": "UI-TP-002",
            "obj_name": "异常OBJ", "fp_name": "异常FP",
            "obj_id": "OBJ-ERR-001", "feature_point_ref": "OBJ-ERR-001-FP-1",
        },
        {  # 4 — Draft (L1 fail: missing s5_ref field — required by L1S6Validator)
            "case_id": "TC-5", "module": "LOG",
            "用例描述": "审计", "功能描述": "日志审计",
            "preconditions": ["ok"], "steps": [{"step_num": 1, "action": "log"}],
            "expected_results": ["已记"], "priority": "P1",
            # NOTE: omit s5_ref to force L1 fail
            "obj_name": "审计OBJ", "fp_name": "审计FP",
            "obj_id": "OBJ-LOG-001", "feature_point_ref": "OBJ-LOG-001-FP-1",
        },
    ]
    return {"test_cases": cases}


def self_test() -> int:
    """Cover the 4 critical paths without touching any disk artifact."""
    import tempfile

    payload = _build_mini_payload()
    xlsx_path: str = ""

    with tempfile.TemporaryDirectory() as tmpdir:
        src = Path(tmpdir) / "tc.json"
        xlsx = Path(tmpdir) / "out.xlsx"
        src.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        summary = normalize_and_export(
            src_json=src,
            xlsx_out=xlsx,
            objs_json=None,
            tps_json=None,
            xlsx_backup=False,  # never backup inside tmpdir
        )

        # Read-back via openpyxl while tmpdir still exists.
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx))
        sheetnames = wb.sheetnames
        assert "测试用例" in sheetnames, sheetnames
        assert "Draft-Rejected附录" in sheetnames, sheetnames

        main_status = wb["测试用例"]
        appendix_status = wb["Draft-Rejected附录"]

        main_rows = list(main_status.iter_rows(min_row=2, values_only=True))
        appendix_rows = list(appendix_status.iter_rows(min_row=2, values_only=True))

        # Locate the 用例状态 column by header (do not hard-code index).
        headers = [cell.value for cell in main_status[1]]
        status_col = headers.index("用例状态")
        appendix_headers = [cell.value for cell in appendix_status[1]]
        appendix_status_col = appendix_headers.index("用例状态")

        main_statuses = [r[status_col] for r in main_rows if r[status_col]]
        appendix_statuses = [r[appendix_status_col] for r in appendix_rows if r[appendix_status_col]]

    assert "Ready" in main_statuses, f"main sheet empty: {main_statuses}"
    assert any(s in {"Draft", "Rejected"} for s in appendix_statuses), (
        f"appendix sheet empty: {appendix_statuses}"
    )
    assert all(s == "Ready" for s in main_statuses), main_statuses
    assert all(s in {"Draft", "Rejected"} for s in appendix_statuses), appendix_statuses

    print("run_normalize_and_export self-test: PASS")
    return 0


def main() -> dict[str, Any]:
    summary = normalize_and_export(
        src_json=DEFAULT_TC_JSON,
        xlsx_out=DEFAULT_XLSX_OUT,
        objs_json=DEFAULT_OBJS_JSON,
        tps_json=DEFAULT_TPS_JSON,
        xlsx_backup=True,
    )

    # === Round 14 F-A: 自动重生成 tc_tp_gap_report（防再次陈旧）===
    try:
        from tc_tp_gap_report import generate_gap_report
        gap_report_path = DEFAULT_S6_DIR / "tc_tp_gap_report.md"
        gap_result = generate_gap_report(
            test_cases_json=DEFAULT_TC_JSON,
            test_points_json=DEFAULT_TPS_JSON,
            xlsx_path=DEFAULT_XLSX_OUT,
            out_path=gap_report_path,
        )
        summary["gap_report"] = gap_result
        print(f"[F-A] gap_report regenerated → {gap_report_path}")
        print(f"      tp={gap_result['tp_count']} json_tc={gap_result['json_tc_count']} "
              f"xlsx_tc={gap_result['xlsx_tc_count']} coverage={gap_result['coverage_rate']:.4f}")
    except Exception as e:
        summary["gap_report_error"] = str(e)
        print(f"[F-A] gap_report regen skipped: {e}")

    print(json.dumps(_summary_for_stdout(summary), ensure_ascii=False, indent=2))
    return summary


def _summary_for_stdout(summary: Mapping[str, Any]) -> dict[str, Any]:
    """Trim the payload to keys humans want to see."""
    writeback = summary["evaluation"]["writeback"]
    # per-case writer uses `transition_count`; legacy bulk writer uses `changed_cases`.
    # Prefer the per-case key, fall back to the bulk one for backward compatibility.
    transitions = writeback.get("transition_count")
    if transitions is None:
        transitions = writeback.get("changed_cases", 0)
    return {
        "src_json": summary["src_json"],
        "xlsx_out": summary["xlsx_out"],
        "total_cases": summary["total_cases"],
        "before": summary["before"],
        "after": summary["after"],
        "l1_passed": summary["evaluation"]["l1_result"]["passed"],
        "l1_errors": summary["evaluation"]["l1_result"]["stats"],
        "l2_passed": summary["evaluation"]["l2_result"]["passed"],
        "l2_total": summary["evaluation"]["l2_result"].get("total", 0),
        "l2_failed_count": summary["evaluation"]["l2_result"].get("failed_count", 0),
        "writeback_target": writeback.get("target_status"),
        "writeback_changed": transitions,
        "note": ("v3.01 test_cases.json untouched; xlsx rewritten in place"
                 if summary.get("writeback_disabled") else "in-place write performed"),
    }


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--self-test":
        raise SystemExit(self_test())
    main()
