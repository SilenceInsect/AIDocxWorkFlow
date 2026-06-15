#!/usr/bin/env python3
"""反馈收集与导入模块。

支持从多种来源导入执行反馈和审查结果：
  1. Excel 文件（含"反馈结果"/"审核结果"/"执行追踪" sheet）
  2. CSV 文件
  3. JSON 文件
  4. test_cases.xlsx 的"执行追踪" sheet

用法：
  python3 -m ai_workflow.feedback_logger \\
      --file /path/to/feedback.xlsx \\
      --version v1.0

  # 或在 Python 中直接调用
  from ai_workflow.feedback_logger import import_feedback
  result = import_feedback("/path/to/feedback.xlsx", version="v1.0")
"""

from __future__ import annotations

import csv, json, sys, argparse
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None   # type: ignore


# ── 标准化列名映射 ──────────────────────────────────────────────────────────
# 支持中文列名和英文列名，自动映射到标准字段

_COLUMN_ALIASES: dict[str, str] = {
    # case_id
    "case_id": "case_id",
    "用例ID": "case_id",
    "用例编号": "case_id",
    "测试用例ID": "case_id",
    "test_case_id": "case_id",
    # result
    "result": "result",
    "执行结果": "result",
    "测试结果": "result",
    "执行状态": "result",
    "test_result": "result",
    "status": "result",
    # actual_result
    "actual_result": "actual_result",
    "实际结果": "actual_result",
    "实际输出": "actual_result",
    "观测结果": "actual_result",
    "actual_output": "actual_result",
    # defect_desc
    "defect_desc": "defect_desc",
    "缺陷描述": "defect_desc",
    "问题描述": "defect_desc",
    "缺陷说明": "defect_desc",
    "bug_description": "defect_desc",
    "issue": "defect_desc",
    # severity
    "severity": "severity",
    "严重程度": "severity",
    "优先级": "severity",
    "缺陷等级": "severity",
    "bug_severity": "severity",
    "priority": "severity",
    # reviewer / audit_by
    "reviewer": "reviewer",
    "审核人": "reviewer",
    "审核者": "reviewer",
    "audit_by": "reviewer",
    "auditor": "reviewer",
    # review_result
    "review_result": "review_result",
    "审核结果": "review_result",
    "审查结果": "review_result",
    "review_status": "review_result",
    "audit_result": "review_result",
    # module
    "module": "module",
    "模块": "module",
    "所属模块": "module",
    # title
    "title": "title",
    "用例描述": "title",
    "用例名称": "title",
    "test_title": "title",
    "test_case_title": "title",
    # execution_time
    "execution_time": "execution_time",
    "执行时间": "execution_time",
    "测试时间": "execution_time",
    "test_time": "execution_time",
    # notes
    "notes": "notes",
    "备注": "notes",
    "comment": "notes",
    "comments": "notes",
    "说明": "notes",
}

# 标准化结果值
_RESULT_MAP: dict[str, str] = {
    "PASS": "PASS",
    "pass": "PASS",
    "通过": "PASS",
    "成功": "PASS",
    "✅": "PASS",
    "true": "PASS",
    "1": "PASS",
    "FAIL": "FAIL",
    "fail": "FAIL",
    "失败": "FAIL",
    "不通过": "FAIL",
    "❌": "FAIL",
    "false": "FAIL",
    "0": "FAIL",
    "BLOCKED": "BLOCKED",
    "blocked": "BLOCKED",
    "阻塞": "BLOCKED",
    "阻断": "BLOCKED",
    "N/A": "N/A",
    "skip": "SKIP",
    "跳过": "SKIP",
    "SKIP": "SKIP",
    "DRAFT": "DRAFT",
    "待执行": "DRAFT",
    "未执行": "DRAFT",
}

_SEVERITY_MAP: dict[str, str] = {
    "CRITICAL": "CRITICAL",
    "critical": "CRITICAL",
    "阻断级": "CRITICAL",
    "P0": "CRITICAL",
    "致命": "CRITICAL",
    "MAJOR": "MAJOR",
    "major": "MAJOR",
    "严重": "MAJOR",
    "P1": "MAJOR",
    "MINOR": "MINOR",
    "minor": "MINOR",
    "一般": "MINOR",
    "P2": "MINOR",
    "LOW": "LOW",
    "low": "LOW",
    "轻微": "LOW",
    "提示": "LOW",
    "P3": "LOW",
}

_REVIEW_RESULT_MAP: dict[str, str] = {
    "PASS": "PASS",
    "pass": "PASS",
    "通过": "PASS",
    "✅": "PASS",
    "合格": "PASS",
    "ACCEPT": "PASS",
    "FAIL": "FAIL",
    "fail": "FAIL",
    "不通过": "FAIL",
    "❌": "FAIL",
    "不合格": "FAIL",
    "REJECT": "FAIL",
    "REVISION": "NEEDS_REVISION",
    "REVISION": "NEEDS_REVISION",
    "revision": "NEEDS_REVISION",
    "需修订": "NEEDS_REVISION",
    "NEEDS_REVISION": "NEEDS_REVISION",
    "N/A": "N/A",
    "DRAFT": "DRAFT",
    "待审核": "DRAFT",
}


# ── 数据模型 ───────────────────────────────────────────────────────────────

@dataclass
class FeedbackEntry:
    case_id: str
    result: str          # PASS / FAIL / BLOCKED / DRAFT
    actual_result: str = ""
    defect_desc: str = ""
    severity: str = ""   # CRITICAL / MAJOR / MINOR / LOW
    module: str = ""
    title: str = ""
    execution_time: str = ""
    notes: str = ""
    source_file: str = ""
    import_time: str = ""


@dataclass
class ReviewEntry:
    case_id: str
    review_result: str    # PASS / FAIL / NEEDS_REVISION / DRAFT
    reviewer: str = ""
    coverage: float = 0.0
    structure_pass_rate: float = 0.0
    overall_pass: bool = False
    notes: str = ""
    source_file: str = ""
    import_time: str = ""


@dataclass
class ImportResult:
    feedback_entries: list[FeedbackEntry] = field(default_factory=list)
    review_entries: list[ReviewEntry] = field(default_factory=list)
    feedback_file: str = ""
    review_file: str = ""
    total_cases: int = 0
    pass_count: int = 0
    fail_count: int = 0
    blocked_count: int = 0
    draft_count: int = 0
    duplicate_count: int = 0
    errors: list[str] = field(default_factory=list)
    saved_to: list[str] = field(default_factory=list)


# ── 核心导入函数 ────────────────────────────────────────────────────────────

def import_feedback(
    feedback_file: str | Path,
    review_file: str | Path | None = None,
    output_dir: str | Path | None = None,
    req_name: str = "游戏道具商城系统",
    version: str = "v1.0",
    merge_duplicates: str = "latest",  # "latest" | "first" | "all"
) -> ImportResult:
    """
    导入反馈/审查结果。

    参数:
        feedback_file: 反馈 Excel/CSV/JSON 文件路径
        review_file:   审查结果 Excel/CSV/JSON 文件路径（可选）
        output_dir:    输出目录（默认 workflow_assets/feedback_logs/）
        req_name:      需求名称
        version:       版本
        merge_duplicates: 重复 case_id 的处理策略

    返回:
        ImportResult: 包含统计和错误信息
    """
    feedback_file = Path(feedback_file)
    result = ImportResult(feedback_file=str(feedback_file))

    if output_dir is None:
        output_dir = Path(__file__).parent.parent / "workflow_assets" / "feedback_logs"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── 读取反馈 ────────────────────────────────────────────────────────
    if feedback_file.exists():
        entries = _read_entries(feedback_file, "feedback")
        result.feedback_entries = _normalize_and_merge(
            entries, merge_duplicates)
        result.total_cases = len(result.feedback_entries)
        for e in result.feedback_entries:
            if e.result == "PASS":  result.pass_count += 1
            elif e.result == "FAIL": result.fail_count += 1
            elif e.result == "BLOCKED": result.blocked_count += 1
            elif e.result == "DRAFT": result.draft_count += 1
    else:
        result.errors.append(f"反馈文件不存在: {feedback_file}")

    # ── 读取审查结果 ────────────────────────────────────────────────────
    if review_file and Path(review_file).exists():
        review_entries = _read_entries(Path(review_file), "review")
        result.review_entries = _normalize_review(review_entries)

    # ── 保存 ─────────────────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if result.feedback_entries:
        fb_path = output_dir / f"feedback_{req_name}_{version}_{ts}.json"
        _save_feedback_json(result.feedback_entries, fb_path)
        result.saved_to.append(str(fb_path))

    if result.review_entries:
        rv_path = output_dir / f"review_{req_name}_{version}_{ts}.json"
        _save_review_json(result.review_entries, rv_path)
        result.saved_to.append(str(rv_path))

    # ── 打印摘要 ────────────────────────────────────────────────────────
    print_import_summary(result)
    return result


def import_from_test_cases_xlsx(
    xlsx_path: str | Path,
    output_dir: str | Path | None = None,
    req_name: str = "游戏道具商城系统",
    version: str = "v1.0",
) -> ImportResult:
    """从 test_cases.xlsx 读取「执行追踪」sheet 作为反馈。"""
    if openpyxl is None:
        result = ImportResult(errors=["openpyxl 未安装，请先 pip install openpyxl"])
        return result

    xlsx_path = Path(xlsx_path)
    result = ImportResult(feedback_file=str(xlsx_path))

    if output_dir is None:
        output_dir = Path(__file__).parent.parent / "workflow_assets" / "feedback_logs"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(str(xlsx_path))
    except Exception as exc:
        result.errors.append(f"无法打开 Excel 文件: {exc}")
        return result

    # 优先读「反馈结果」，其次「执行追踪」
    sheet_names = wb.sheetnames
    target_sheet = None
    for preferred in ["反馈结果", "执行追踪", "执行记录"]:
        if preferred in sheet_names:
            target_sheet = preferred
            break

    if target_sheet is None:
        result.errors.append(f"未找到「反馈结果」或「执行追踪」sheet，可用: {sheet_names}")
        return result

    ws = wb[target_sheet]
    headers = [str(h).strip() if h else "" for h in next(ws.iter_rows(values_only=True))]
    entries_raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        entry = {headers[i]: str(v).strip() if v is not None else "" for i, v in enumerate(row)}
        entries_raw.append(entry)

    result.feedback_entries = _normalize_and_merge(entries_raw, "latest")
    result.total_cases = len(result.feedback_entries)
    for e in result.feedback_entries:
        if e.result == "PASS":  result.pass_count += 1
        elif e.result == "FAIL": result.fail_count += 1
        elif e.result == "BLOCKED": result.blocked_count += 1
        elif e.result == "DRAFT": result.draft_count += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fb_path = output_dir / f"feedback_{req_name}_{version}_{ts}.json"
    _save_feedback_json(result.feedback_entries, fb_path)
    result.saved_to.append(str(fb_path))

    print_import_summary(result)
    return result


# ── 内部函数 ───────────────────────────────────────────────────────────────

def _read_entries(fpath: Path, entry_type: str) -> list[dict]:
    """根据文件类型读取原始条目列表。"""
    suffix = fpath.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _read_excel(fpath, entry_type)
    elif suffix == ".csv":
        return _read_csv(fpath)
    elif suffix == ".json":
        return _read_json(fpath)
    return []


def _read_excel(fpath: Path, entry_type: str) -> list[dict]:
    """读取 Excel，自动探测「反馈结果」/「审核结果」sheet。"""
    if openpyxl is None:
        return []
    try:
        wb = openpyxl.load_workbook(str(fpath))
    except Exception:
        return []

    sheets_to_try = []
    if entry_type == "feedback":
        sheets_to_try = ["反馈结果", "执行追踪", "执行记录", "用例详情"]
    elif entry_type == "review":
        sheets_to_try = ["审核结果", "审查结果", "review"]

    target = None
    for name in sheets_to_try:
        if name in wb.sheetnames:
            target = name
            break

    if target is None:
        # 取第一个有数据的 sheet
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row > 1:
                target = name
                break

    if target is None:
        return []

    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    entries = []
    for row in rows[1:]:
        if not any(v for v in row):
            continue
        entry = {}
        for i, val in enumerate(row):
            if i < len(headers):
                entry[headers[i]] = str(val).strip() if val is not None else ""
        entries.append(entry)
    return entries


def _read_csv(fpath: Path) -> list[dict]:
    entries = []
    try:
        with fpath.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                entries.append({k.strip(): str(v).strip() if v else "" for k, v in row.items()})
    except Exception:
        try:
            with fpath.open(encoding="gbk", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    entries.append({k.strip(): str(v).strip() if v else "" for k, v in row.items()})
        except Exception:
            pass
    return entries


def _read_json(fpath: Path) -> list[dict]:
    try:
        with fpath.open(encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            # 尝试从常见字段中提取列表
            for key in ("feedback", "entries", "data", "results"):
                if isinstance(data.get(key), list):
                    return data[key]
            return [data]
    except Exception:
        pass
    return []


def _normalize_alias(raw_key: str) -> str:
    return _COLUMN_ALIASES.get(raw_key.strip(), raw_key.strip())


def _map_value(val: str, mapping: dict[str, str]) -> str:
    return mapping.get(val.strip(), val.strip())


def _normalize_and_merge(entries_raw: list[dict], merge_strategy: str) -> list[FeedbackEntry]:
    """将原始条目标准化并合并重复的 case_id。"""
    normalized = []
    by_id: dict[str, FeedbackEntry] = {}

    for raw in entries_raw:
        if not raw:
            continue

        # 标准化键名
        std = {_normalize_alias(k): v for k, v in raw.items()}
        cid = std.get("case_id", "").strip()
        if not cid:
            continue

        res = _map_value(std.get("result", "DRAFT"), _RESULT_MAP) or "DRAFT"

        entry = FeedbackEntry(
            case_id=cid,
            result=res,
            actual_result=std.get("actual_result", ""),
            defect_desc=std.get("defect_desc", ""),
            severity=_map_value(std.get("severity", ""), _SEVERITY_MAP),
            module=std.get("module", ""),
            title=std.get("title", ""),
            execution_time=std.get("execution_time", ""),
            notes=std.get("notes", ""),
            source_file="",
            import_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

        if cid not in by_id:
            by_id[cid] = entry
        else:
            existing = by_id[cid]
            if merge_strategy == "latest":
                # 用最新的非空字段覆盖
                for attr in ("result", "actual_result", "defect_desc", "severity",
                              "module", "title", "execution_time", "notes"):
                    new_val = getattr(entry, attr)
                    if new_val:
                        setattr(existing, attr, new_val)
            elif merge_strategy == "all":
                # 保留所有结果（多个记录）
                normalized.append(existing)
                by_id[cid] = entry

    result = list(by_id.values())
    if merge_strategy == "all":
        result = [e for e in result if e not in normalized]
        result = list(by_id.values())
    return result


def _normalize_review(entries_raw: list[dict]) -> list[ReviewEntry]:
    normalized = []
    for raw in entries_raw:
        if not raw:
            continue
        std = {_normalize_alias(k): v for k, v in raw.items()}
        cid = std.get("case_id", "").strip()
        if not cid:
            continue
        rev_res = _map_value(std.get("review_result", "DRAFT"), _REVIEW_RESULT_MAP) or "DRAFT"
        cov_str = std.get("coverage", "0")
        try:
            cov = float(cov_str.replace("%", "")) / 100 if isinstance(cov_str, str) else float(cov_str)
        except (ValueError, AttributeError):
            cov = 0.0

        normalized.append(ReviewEntry(
            case_id=cid,
            review_result=rev_res,
            reviewer=std.get("reviewer", ""),
            coverage=round(cov, 3),
            structure_pass_rate=float(std.get("structure_pass_rate", 0)),
            overall_pass=rev_res == "PASS",
            notes=std.get("notes", ""),
            source_file="",
            import_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))
    return normalized


def _save_feedback_json(entries: list[FeedbackEntry], path: Path) -> None:
    data = [asdict(e) for e in entries]
    with path.open("w", encoding="utf-8") as f:
        json.dump({"version": "v1.0", "entries": data, "count": len(data)},
                  f, ensure_ascii=False, indent=2)
    print(f"[FeedbackLogger] ✅ 已保存 {len(data)} 条反馈 → {path}")


def _save_review_json(entries: list[ReviewEntry], path: Path) -> None:
    data = [asdict(e) for e in entries]
    with path.open("w", encoding="utf-8") as f:
        json.dump({"version": "v1.0", "entries": data, "count": len(data)},
                  f, ensure_ascii=False, indent=2)
    print(f"[FeedbackLogger] ✅ 已保存 {len(data)} 条审核 → {path}")


def print_import_summary(result: ImportResult) -> None:
    print(f"\n{'='*55}")
    print(f"  反馈导入结果")
    print(f"{'='*55}")
    print(f"  来源文件：{result.feedback_file or '-'}")
    if result.review_file:
        print(f"  审核来源：{result.review_file}")
    print(f"  用例总数：{result.total_cases}")
    print(f"  ✅ PASS:    {result.pass_count:>5}")
    print(f"  ❌ FAIL:    {result.fail_count:>5}")
    print(f"  ⏸ BLOCKED: {result.blocked_count:>5}")
    print(f"  📝 DRAFT:   {result.draft_count:>5}")
    if result.duplicate_count:
        print(f"  合并重复：{result.duplicate_count} 个 case_id")
    if result.errors:
        print(f"  ⚠️ 错误：")
        for err in result.errors:
            print(f"    - {err}")
    print(f"  保存至：")
    for p in result.saved_to:
        print(f"    - {p}")
    print(f"{'='*55}\n")


# ── CLI ────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="AIDocxWorkFlow 反馈导入工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  # 从 Excel 导入（自动探测反馈结果 sheet）
  python3 -m ai_workflow.feedback_logger --file feedback.xlsx

  # 从 test_cases.xlsx 的执行追踪 sheet 导入
  python3 -m ai_workflow.feedback_logger --file test_cases.xlsx

  # 同时提供反馈文件和审查结果文件
  python3 -m ai_workflow.feedback_logger \\
      --file feedback.xlsx \\
      --review review.xlsx \\
      --version v1.0

  # 指定输出目录
  python3 -m ai_workflow.feedback_logger --file feedback.xlsx \\
      --output /path/to/custom/feedback_logs/
        """,
    )
    parser.add_argument("--file",   required=True, help="反馈结果文件（Excel/CSV/JSON）")
    parser.add_argument("--review", help="审查结果文件（可选，Excel/CSV/JSON）")
    parser.add_argument("--output", help="输出目录（默认: workflow_assets/feedback_logs/）")
    parser.add_argument("--version", default="v1.0", help="版本号（默认 v1.0）")
    parser.add_argument("--req-name", default="游戏道具商城系统",
                        help="需求名称（默认: 游戏道具商城系统）")

    args = parser.parse_args()
    import_feedback(
        feedback_file=args.file,
        review_file=args.review,
        output_dir=args.output,
        version=args.version,
        req_name=args.req_name,
    )


if __name__ == "__main__":
    main()
